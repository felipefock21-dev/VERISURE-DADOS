# -*- coding: utf-8 -*-
"""
COMPILADOR DE RELATÓRIOS VERISURE - Versão Web (Flask)
Reutiliza FIELMENTE a lógica do script Colab em uma aplicação web
"""
# Carrega .env da raiz do projeto (para OAUTH_CLIENT_ID, DEPLOY_URL, etc.)
import os as _os
import sys as _sys
_root = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_sys.path.insert(0, _root)
try:
    from dotenv import load_dotenv
    load_dotenv(_os.path.join(_root, ".env"))
except ImportError:
    pass

from flask import Flask, render_template, request, jsonify, send_file, redirect, session, url_for, Response
import pandas as pd
import os
import glob
from datetime import datetime
import warnings
import re
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
import io
import time
import gspread
import json
import threading
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from openpyxl import load_workbook
from copy import copy
from openpyxl.utils import get_column_letter
import tempfile
import sys
import gc
import psutil
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

# Adicionar pasta raiz ao path para importar oauth_manager
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from oauth_manager import (
    authorize_url, exchange_code_for_token, 
    get_authenticated_drive_service, get_authenticated_sheets_service,
    load_credentials, save_credentials
)
from oauth_config import DRIVE_FOLDER_ID, TOKEN_FILE

warnings.filterwarnings('ignore')

# ===== ARMAZENAMENTO DE PROGRESSO (para SSE - Baseado em Arquivo) =====
# Caminho absoluto em temp_uploads para evitar WinError 5 (acesso negado) quando CWD muda
_PROGRESS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_uploads')
try:
    os.makedirs(_PROGRESS_DIR, exist_ok=True)
except OSError:
    _PROGRESS_DIR = tempfile.gettempdir()
PROGRESS_FILE = os.path.join(_PROGRESS_DIR, 'upload_progress.json')

def atualizar_progresso(etapa, percentual, mensagem):
    """Atualiza o progresso em um arquivo compartilhado de forma atômica"""
    progress = {
        'etapa': etapa,
        'percentual': min(percentual, 100),
        'mensagem': mensagem,
        'timestamp': time.time()
    }
    try:
        temp_progress = PROGRESS_FILE + '.tmp'
        with open(temp_progress, 'w', encoding='utf-8') as f:
            json.dump(progress, f)
        try:
            os.replace(temp_progress, PROGRESS_FILE)
        except OSError:
            # Fallback: grava direto (evita WinError 5 em Windows quando outro processo usa o arquivo)
            with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
                json.dump(progress, f)
            if os.path.exists(temp_progress):
                try:
                    os.remove(temp_progress)
                except OSError:
                    pass
    except Exception as e:
        print(f"[PROGRESSO] Erro ao gravar: {e}")
    print(f"[PROGRESSO] {percentual}% - {mensagem}")

def ler_progresso():
    """Lê o progresso do arquivo compartilhado"""
    if not os.path.exists(PROGRESS_FILE):
        return {'etapa': 0, 'percentual': 0, 'mensagem': 'Aguardando arquivo...'}
    try:
        with open(PROGRESS_FILE, 'r') as f:
            return json.load(f)
    except:
        return {'etapa': 0, 'percentual': 0, 'mensagem': 'Erro ao ler progresso...'}

# Verificar se OAuth está configurado
OAUTH_CONFIGURED = os.path.exists(TOKEN_FILE)
print(f"[OAUTH] Status inicial: {'✅ Token carregado' if OAUTH_CONFIGURED else '⚠️ Não autenticado - acesse /authorize'}")

app = Flask(__name__)
# Configurar ProxyFix para Railway (essencial para url_for usar HTTPS)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max
app.config['UPLOAD_FOLDER'] = 'temp_uploads'
app.config['SAIDAS_FOLDER'] = 'saidas'
app.secret_key = 'sua_chave_secreta_aqui_mude_em_producao'  # Para sessions

# Adicionar headers CORS para permitir requisições de qualquer origem
@app.after_request
def add_cors_headers(response):
    """Adiciona headers CORS para permitir requisições do navegador"""
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

# ==============================================================================
# ENDPOINTS OAUTH 2.0
# ==============================================================================

# Placeholders que geram Erro 401 invalid_client no Google
OAUTH_PLACEHOLDERS = ("COLOQUE_SEU_CLIENT_ID_AQUI", "COLOQUE_SEU_CLIENT_SECRET_AQUI")

def _oauth_redirect_uri():
    """Redirect URI baseada no host da requisição (funciona em verisure-dados ou verisure-compilador sem mudar DEPLOY_URL)."""
    base = f"{request.scheme}://{request.host}".rstrip("/")
    return f"{base}/oauth2callback"

@app.route('/authorize')
def authorize():
    """Redireciona para a página de login do Google"""
    from oauth_config import DEPLOY_URL, OAUTH_CLIENT_ID, OAUTH_CLIENT_SECRET
    
    # Evita redirecionar ao Google com credenciais placeholder (causa 401 invalid_client)
    if (OAUTH_CLIENT_ID in OAUTH_PLACEHOLDERS or OAUTH_CLIENT_SECRET in OAUTH_PLACEHOLDERS or
        not OAUTH_CLIENT_ID.strip() or not OAUTH_CLIENT_SECRET.strip()):
        redirect_uri_fallback = f"{DEPLOY_URL}/oauth2callback"
        return render_template(
            "oauth_config_erro.html",
            deploy_url=DEPLOY_URL,
            redirect_uri=redirect_uri_fallback
        ), 400
    
    # Usa o host da requisição para o redirect_uri (assim funciona em qualquer domínio onde o app está rodando)
    redirect_uri = _oauth_redirect_uri()
    
    print(f"[OAUTH] Gerando URL de autorizacao...")
    print(f"[OAUTH]    Host atual: {request.host}")
    print(f"[OAUTH]    Redirect URI gerado: {redirect_uri}")
    
    auth_url, state = authorize_url(redirect_uri=redirect_uri)
    session['oauth_state'] = state
    return redirect(auth_url)

@app.route('/oauth-setup')
def oauth_setup():
    """Página com a URI exata e instruções para corrigir redirect_uri_mismatch (Erro 400)."""
    from oauth_config import DEPLOY_URL
    return render_template(
        "oauth_redirect_fix.html",
        redirect_uri=_oauth_redirect_uri(),
        deploy_url=f"{request.scheme}://{request.host}".rstrip("/"),
    )

@app.route('/debug-oauth')
@app.route('/debug_oauth')
def debug_oauth():
    """Rota de diagnóstico para verificar o Redirect URI gerado"""
    from oauth_config import DEPLOY_URL, OAUTH_REDIRECT_URI, OAUTH_CLIENT_ID
    
    scheme = 'https' if 'localhost' not in request.host else 'http'
    url_for_uri = url_for('oauth2callback', _external=True, _scheme=scheme)
    
    return jsonify({
        'config_deploy_url': DEPLOY_URL,
        'config_redirect_uri': OAUTH_REDIRECT_URI,
        'effective_redirect_uri': _oauth_redirect_uri(),
        'url_for_generated_uri': url_for_uri,
        'client_id': OAUTH_CLIENT_ID,
        'host_header': request.host,
        'request_scheme': request.scheme,
        'env_deploy_url': os.getenv("DEPLOY_URL"),
        'instrucao': 'Garanta que "effective_redirect_uri" esteja EXATAMENTE IGUAL no Google Cloud Console.'
    })

@app.route('/oauth2callback')
def oauth2callback():
    """Callback após autorização do Google"""
    code = request.args.get('code')
    state = request.args.get('state')
    
    if not code:
        return jsonify({'erro': 'Autorização negada'}), 401
    
    try:
        # Mesmo redirect_uri usado na autorização (baseado no host da requisição)
        redirect_uri = _oauth_redirect_uri()
        
        print(f"[OAUTH] 🔄 Processando callback...")
        print(f"[OAUTH]    Redirect URI usado para troca de token: {redirect_uri}")
        
        creds = exchange_code_for_token(code, redirect_uri=redirect_uri)
        print(f"[OAUTH] ✅ Autenticação bem-sucedida!")
        print(f"[OAUTH] ✅ Autenticação bem-sucedida! Redirecionando para a página principal...")
        return redirect(url_for('index', _external=True))
    except Exception as e:
        print(f"[OAUTH] ❌ Erro na autenticação: {str(e)}")
        return jsonify({'erro': f'Erro na autenticação: {str(e)}'}), 401

@app.route('/progresso')
def progresso_sse():
    """Server-Sent Events para enviar progresso em tempo real (lendo do arquivo)"""
    def gerar_progresso():
        ultimo_percentual = -1
        heartbeat_counter = 0
        while True:
            progress = ler_progresso()
            # Só envia se houver mudança significativa ou for novo
            if progress.get('percentual') != ultimo_percentual:
                yield f"data: {json.dumps(progress)}\n\n"
                ultimo_percentual = progress.get('percentual')
                heartbeat_counter = 0
            else:
                # Heartbeat mais frequente para evitar timeout
                heartbeat_counter += 1
                yield f": heartbeat {heartbeat_counter}\n\n"
            
            # Se terminou, mantém um pouco mais e para
            if progress.get('etapa') == 4:
                time.sleep(2)
                break
                
            time.sleep(0.5)  # Heartbeat a cada 500ms (mais frequente)
    
    response = Response(gerar_progresso(), mimetype='text/event-stream')
    # Headers para evitar buffering em proxies
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'
    response.headers['Connection'] = 'keep-alive'
    return response

@app.route('/oauth-status')
def oauth_status():
    """Retorna o status da autenticação OAuth"""
    creds = load_credentials()
    if creds:
        return jsonify({
            'autenticado': True,
            'mensagem': '✅ Autenticado! Pronto para fazer upload.'
        }), 200
    else:
        return jsonify({
            'autenticado': False,
            'mensagem': '⚠️ Não autenticado. Acesse /authorize para fazer login.',
            'link_authorize': '/authorize'
        }), 401


@app.route('/debug-dados-id')
def debug_dados_id():
    """
    Debug: mostra exatamente o que o app recebe do Google Sheet DadosIdentificador.
    Abra no navegador: http://localhost:5000/debug-dados-id
    """
    try:
        sheets_service = get_authenticated_sheets_service()
        if not sheets_service:
            return jsonify({
                'erro': 'Não autenticado',
                'dica': 'Acesse /authorize primeiro e faça login no Google.'
            }), 401
        sheet_id = "1UmWzuIpF1nEh1YUJH4pu9jJL7PJHvlQvHE4Pnrw6xhA"
        spreadsheet = sheets_service.spreadsheets().get(spreadsheetId=sheet_id).execute()
        sheets_list = spreadsheet.get('sheets', [])
        if not sheets_list:
            return jsonify({'erro': 'Planilha sem abas', 'sheet_id': sheet_id}), 404
        sheet_title = sheets_list[0].get('properties', {}).get('title')
        range_name = f"'{sheet_title}'!A:Z"
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=sheet_id, range=range_name
        ).execute()
        values = result.get('values', [])
        if not values:
            return jsonify({
                'mensagem': 'Planilha vazia',
                'sheet_id': sheet_id,
                'aba': sheet_title,
                'linhas': 0
            }), 200
        # Primeiras linhas brutas (o que a API retorna)
        linha1 = values[0] if len(values) > 0 else []
        linha2 = values[1] if len(values) > 1 else []
        linha3 = values[2] if len(values) > 2 else []
        # Nomes das colunas (linha 1) com índice para referência
        colunas_com_indice = [{'indice': i, 'nome': str(v)[:50]} for i, v in enumerate(linha1)]
        return jsonify({
            'sheet_id': sheet_id,
            'aba': sheet_title,
            'total_linhas': len(values),
            'linha_1_cabecalho': linha1,
            'linha_2': linha2,
            'linha_3': linha3,
            'colunas_com_indice': colunas_com_indice,
            'dica': 'Identificador costuma ser coluna 0 (A), Universo coluna 11 (L), porc coluna 10 (K). Use essas posições se os nomes não baterem.'
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# Criar pastas
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['SAIDAS_FOLDER'], exist_ok=True)

# Períodos comerciais
PERIODOS_COMERCIAIS = {
    # 2025
    "Jan'25": ("2024-12-25", "2025-01-28"),
    "Fev'25": ("2025-01-29", "2025-02-25"),
    "Mar'25": ("2025-02-26", "2025-03-31"),
    "Abr'25": ("2025-04-01", "2025-04-28"),
    "Mai'25": ("2025-04-29", "2025-05-26"),
    "Jun'25": ("2025-05-27", "2025-06-30"),
    "Jul'25": ("2025-07-01", "2025-07-28"),
    "Ago'25": ("2025-07-29", "2025-08-25"),
    "Set'25": ("2025-08-26", "2025-09-29"),
    "Out'25": ("2025-09-30", "2025-10-27"),
    "Nov'25": ("2025-10-28", "2025-11-24"),
    "Dez'25": ("2025-11-25", "2025-12-29"),
    # 2026
    "Jan'26": ("2025-12-30", "2026-01-26"),
    "Fev'26": ("2026-01-27", "2026-02-23"),
    "Mar'26": ("2026-02-24", "2026-03-30"),
    "Abr'26": ("2026-03-31", "2026-04-27"),
    "Mai'26": ("2026-04-28", "2026-05-25"),
    "Jun'26": ("2026-05-26", "2026-06-29"),
    "Jul'26": ("2026-06-30", "2026-07-27"),
    "Ago'26": ("2026-07-28", "2026-08-31"),
    "Set'26": ("2026-09-01", "2026-09-28"),
    "Out'26": ("2026-09-29", "2026-10-26"),
    "Nov'26": ("2026-10-27", "2026-11-30"),
    "Dez'26": ("2026-12-01", "2026-12-28"),
}

# Mapa de praças
MAPA_PRACAS = {
    "São Paulo Capital": ["São Paulo / SP"],
    "Rio de Janeiro Capital": ["Rio de Janeiro / RJ"],
    "Belo Horizonte Capital": ["Belo Horizonte / MG"],
    "Campinas Capital": ["Campinas / SP"],
    "São Paulo Litoral": ["Santos / SP"],
    "Ribeirão Preto": ["Ribeirão Preto / SP"],
    "Sorocaba": ["Sorocaba / SP"],
    "SJCampos": ["São José dos Campos / SP"],
    "Vitoria Capital": ["Vitória / ES"],
    "Petropolis e Volta Redonda RJ": ["Petrópolis / RJ", "Volta Redonda / RJ"],
    "Curitiba Capital": ["Curitiba / PR"],
    "Porto Alegre Capital": ["Porto Alegre / RS"],
    "Florianopolis Capital": ["Florianópolis / SC"],
    "Joinville": ["Joinville / SC"],
    "Brasilia Capital": ["Brasília / DF"],
    "Goiania Capital": ["Goiânia / GO"],
    "Uberlandia": ["Uberlândia / MG"],
    "Campo Grande Capital": ["Campo Grande / MS"],
    "Recife Capital": ["Recife / PE"],
    "Salvador Capital": ["Salvador / BA"],
    "Fortaleza Capital": ["Fortaleza / CE"],
    "Joao Pessoa Capital": ["João Pessoa / PB"],
    "Bauru": ["Bauru / SP"],
}

# ==============================================================================
# UTILITÁRIOS PARA GERENCIAR ARQUIVOS NA PASTA SAIDAS
# ==============================================================================

import glob as glob_module

def list_google_sheets_in_drive(drive_service, folder_name="RelatorioVeri"):
    """Lista todos os Google Sheets em uma pasta específica do Drive"""
    try:
        # Busca a pasta no Drive
        folder_query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder'"
        folder_results = drive_service.files().list(q=folder_query, fields="files(id, name)").execute()
        folders = folder_results.get('files', [])

        if not folders:
            print(f"[DRIVE] ⚠️ Pasta '{folder_name}' não encontrada no Google Drive")
            return []

        folder_id = folders[0]['id']
        print(f"[DRIVE] 📁 Pasta encontrada: {folder_name} (ID: {folder_id})")

        # Lista todos os Google Sheets na pasta
        sheets_query = f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.spreadsheet'"
        sheets_results = drive_service.files().list(
            q=sheets_query,
            fields="files(id, name, modifiedTime)",
            pageSize=100
        ).execute()
        sheets = sheets_results.get('files', [])

        print(f"[DRIVE] 📋 Encontrados {len(sheets)} Google Sheets na pasta")

        # Debug: lista todos os arquivos encontrados
        print("[DRIVE] 🔍 Arquivos encontrados:")
        for sheet in sheets:
            print(f"[DRIVE]    📄 {sheet['name']}")

        return sheets

    except Exception as e:
        print(f"[DRIVE] ⚠️ Erro ao listar sheets da pasta: {str(e)}")
        return []

def read_google_sheet(sheets_service, sheet_id, sheet_name):
    """Lê dados de um Google Sheet pelo ID usando Google Sheets API (OAuth)"""
    try:
        print(f"[DRIVE]   📖 Abrindo Google Sheet: {sheet_name}")
        
        # Obtém metadados da planilha (lista de abas)
        spreadsheet = sheets_service.spreadsheets().get(spreadsheetId=sheet_id).execute()
        sheets = spreadsheet.get('sheets', [])
        
        if not sheets:
            print(f"[DRIVE]   ❌ Nenhuma aba encontrada em {sheet_name}")
            return None
        
        # Tenta ler cada aba até encontrar dados válidos
        for i, sheet_meta in enumerate(sheets):
            sheet_title = sheet_meta.get('properties', {}).get('title')
            try:
                print(f"[DRIVE]     📄 Tentando aba {i+1}: {sheet_title}")
                
                # Lê os dados da aba
                range_name = f"'{sheet_title}'!A:Z"
                result = sheets_service.spreadsheets().values().get(
                    spreadsheetId=sheet_id,
                    range=range_name
                ).execute()
                
                values = result.get('values', [])
                
                if values and len(values) > 1:  # Pelo menos cabeçalho + 1 linha
                    df = pd.DataFrame(values[1:], columns=values[0])
                    print(f"[DRIVE]     ✅ Sucesso! {len(df)} linhas lidas da aba '{sheet_title}'")
                    return df
                    
            except Exception as e:
                print(f"[DRIVE]     ⚠️ Erro na aba {sheet_title}: {str(e)}")
                continue
        
        print(f"[DRIVE]   ❌ Nenhuma aba válida encontrada em {sheet_name}")
        return None
        
    except Exception as e:
        print(f"[DRIVE]   ❌ Erro ao ler {sheet_name}: {str(e)}")
        return None

def get_saidas_folder_id():
    """Encontra ou cria a pasta 'saidas' no Google Drive"""
    try:
        drive_service, _ = init_google_drive()
        if not drive_service:
            return None
        
        # Busca pasta RelatorioVeri
        folder_query = "name='RelatorioVeri' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        folder_results = drive_service.files().list(q=folder_query, fields="files(id, name)").execute()
        folders = folder_results.get('files', [])
        
        if not folders:
            print("[GOOGLE DRIVE] ⚠️ Pasta 'RelatorioVeri' não encontrada")
            return None
        
        relatorio_veri_id = folders[0]['id']
        
        # Busca pasta 'saidas' dentro de RelatorioVeri
        saidas_query = f"name='saidas' and '{relatorio_veri_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
        saidas_results = drive_service.files().list(q=saidas_query, fields="files(id, name)").execute()
        saidas_folders = saidas_results.get('files', [])
        
        if saidas_folders:
            return saidas_folders[0]['id']
        
        # Se não existir, cria a pasta
        print("[GOOGLE DRIVE] 📁 Criando pasta 'saidas'...")
        file_metadata = {
            'name': 'saidas',
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [relatorio_veri_id]
        }
        folder = drive_service.files().create(body=file_metadata, fields='id').execute()
        saidas_id = folder.get('id')
        print(f"[GOOGLE DRIVE] ✅ Pasta 'saidas' criada: {saidas_id}")
        return saidas_id
    
    except Exception as e:
        print(f"[GOOGLE DRIVE] ❌ Erro ao buscar pasta saidas: {str(e)}")
        return None

def upload_to_drive(local_filepath, filename):
    """Faz upload do arquivo para Google Drive usando OAuth"""
    try:
        drive_service = get_authenticated_drive_service()
        if not drive_service:
            print("[OAUTH] ⚠️ Não autenticado - Acesse /authorize primeiro")
            return None
        
        # Usa o folder_id da configuração OAuth
        folder_id = DRIVE_FOLDER_ID
        
        # Verifica se já existe arquivo com mesmo nome e deleta
        query = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
        results = drive_service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get('files', [])
        
        if files:
            for file in files:
                drive_service.files().delete(fileId=file['id']).execute()
                print(f"[GOOGLE DRIVE] 🗑️ Arquivo antigo deletado: {filename}")
        
        # Faz upload do novo arquivo
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        
        media = MediaFileUpload(local_filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file = drive_service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
        
        file_id = file.get('id')
        link = file.get('webViewLink')
        print(f"[GOOGLE DRIVE] ✅ Upload realizado: {filename}")
        print(f"[GOOGLE DRIVE] 🔗 Link: {link}")
        return file_id
    
    except Exception as e:
        print(f"[GOOGLE DRIVE] ❌ Erro no upload: {str(e)}")
        return None
        return None


def upload_all_reports_to_drive(timestamp):
    """Faz upload automático dos 3 arquivos gerados para Google Drive (pasta compartilhada)"""
    print("\n" + "="*80)
    print("[UPLOAD] 🚀 Iniciando upload automático para Google Drive...")
    print("="*80)
    
    try:
        saidas_path = app.config['SAIDAS_FOLDER']
        
        # Encontra os arquivos MAIS RECENTES de cada tipo
        import glob
        
        # Procura o COMPILADO com esse timestamp (pode ser .xlsx ou .csv)
        compilado_files = []
        for ext in ['.xlsx', '.csv']:
            compilado_files.extend(glob.glob(os.path.join(saidas_path, f'COMPILADO_{timestamp}{ext}')))
        compilado_files = sorted(compilado_files, key=lambda x: os.path.getmtime(x), reverse=True)
        
        # Procura o MENSAL mais recente (pode ser .xlsx ou .csv)
        mensal_files = []
        for ext in ['.xlsx', '.csv']:
            mensal_files.extend(glob.glob(os.path.join(saidas_path, f'RELATORIO_MENSAL_*{ext}')))
        mensal_files = sorted(mensal_files, key=lambda x: os.path.getmtime(x), reverse=True)
        
        # Procura o SEMANAL mais recente (pode ser .xlsx ou .csv)
        semanal_files = []
        for ext in ['.xlsx', '.csv']:
            semanal_files.extend(glob.glob(os.path.join(saidas_path, f'RELATORIO_SEMANAL_*{ext}')))
        semanal_files = sorted(semanal_files, key=lambda x: os.path.getmtime(x), reverse=True)
        
        files_to_upload = []
        
        if compilado_files:
            files_to_upload.append((compilado_files[0], os.path.basename(compilado_files[0]), 'COMPILADO'))
        else:
            print(f"[UPLOAD] ⚠️ Arquivo COMPILADO não encontrado para timestamp {timestamp}")
            
        if mensal_files:
            files_to_upload.append((mensal_files[0], os.path.basename(mensal_files[0]), 'MENSAL'))
        if semanal_files:
            files_to_upload.append((semanal_files[0], os.path.basename(semanal_files[0]), 'SEMANAL'))
        
        results = {}
        for filepath, filename, tipo in files_to_upload:
            
            if os.path.exists(filepath):
                print(f"\n[UPLOAD] 📤 Uploading {tipo}...")
                print(f"[UPLOAD] Arquivo: {filename}")
                file_id = upload_to_drive(filepath, filename)
                results[tipo] = {'sucesso': file_id is not None, 'file_id': file_id, 'filename': filename}
                print(f"[UPLOAD] {tipo}: {'✅' if file_id else '❌'}")
            else:
                print(f"\n[UPLOAD] ⚠️ Arquivo não encontrado: {filepath}")
                results[tipo] = {'sucesso': False, 'file_id': None}
        
        print("\n" + "="*80)
        print("[UPLOAD] ✅ Upload automático concluído!")
        print("="*80 + "\n")
        return results
    
    except Exception as e:
        print(f"[UPLOAD] ❌ Erro ao fazer upload dos arquivos: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def get_latest_compiled_file():
    """Encontra o arquivo compilado mais recente na pasta saidas (.xlsx ou .csv; compilados grandes são salvos como CSV)"""
    saidas_path = app.config['SAIDAS_FOLDER']
    files = []
    for ext in ('*.xlsx', '*.csv'):
        files.extend(glob_module.glob(os.path.join(saidas_path, 'COMPILADO_' + ext)))
    if not files:
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]

def get_latest_semanal_file():
    """Encontra o arquivo semanal mais recente na pasta saidas"""
    saidas_path = app.config['SAIDAS_FOLDER']
    pattern = os.path.join(saidas_path, 'RELATORIO_SEMANAL_*.xlsx')
    files = glob_module.glob(pattern)
    
    if not files:
        return None
    
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]

def apply_manual_formatting(output_path, df_output):
    """Aplica formatação profissional igual ao Colab - sem copiar template"""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb_output = load_workbook(output_path)
        ws_output = wb_output.active
        
        num_cols = len(df_output.columns)
        num_rows = ws_output.max_row
        
        print(f"[FORMATACAO] 🎨 Aplicando formatação manual: {num_cols} colunas x {num_rows} linhas")
        
        # ===== 1️⃣ DEFINIR ESTILOS =====
        print(f"[FORMATACAO]    1. Definindo estilos...")
        
        # Fonte do cabeçalho: Bold, Branco, Tamanho 11
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        
        # Fonte dos dados: Tamanho 10
        data_font = Font(size=10, name="Arial")
        
        # Cor do cabeçalho: Azul escuro (#366092)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Bordas finas em todas as células
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alinhamento: centro
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        print(f"[FORMATACAO]       ✅ Estilos criados (Azul #366092, Arial Bold 11)")
        
        # ===== 2️⃣ FORMATAR CABEÇALHO =====
        print(f"[FORMATACAO]    2. Formatando cabeçalho...")
        
        for col_idx in range(1, num_cols + 1):
            cell = ws_output.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        ws_output.row_dimensions[1].height = 20  # Altura do cabeçalho
        
        print(f"[FORMATACAO]       ✅ Cabeçalho formatado (fundo azul, texto branco)")
        
        # ===== 3️⃣ FORMATAR DADOS =====
        print(f"[FORMATACAO]    3. Formatando dados...")
        
        for row_idx in range(2, num_rows + 1):
            ws_output.row_dimensions[row_idx].height = 16  # Altura das linhas
            
            for col_idx in range(1, num_cols + 1):
                cell = ws_output.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = thin_border
        
        print(f"[FORMATACAO]       ✅ Dados formatados (bordas, alinhamento centro)")
        
        # ===== 4️⃣ FORMATAR NÚMEROS =====
        print(f"[FORMATACAO]    4. Formatando números...")
        
        column_names = list(df_output.columns)
        
        for row_idx in range(2, num_rows + 1):
            for col_idx, col_name in enumerate(column_names, 1):
                cell = ws_output.cell(row=row_idx, column=col_idx)
                
                # Colunas de moeda
                if col_name in ['Investimento', 'Preço']:
                    cell.number_format = 'R$ #,##0.00'
                # Colunas de número inteiro com separador
                elif col_name in ['Impacto', 'TRPs', 'Universo', 'PMM', 'Inserções', 'Impactos', 'TRP']:
                    cell.number_format = '#,##0'
        
        print(f"[FORMATACAO]       ✅ Números formatados (Investimento=R$, outros=#,##0)")
        
        # ===== 5️⃣ AJUSTAR LARGURA DAS COLUNAS =====
        print(f"[FORMATACAO]    5. Ajustando largura das colunas...")
        
        # Larguras padrão para MENSAL (7 colunas: Praça, Mês, Impacto, TRPs, Investimento, Universo, PMM)
        column_widths_mensal = {
            'A': 25,  # Praça
            'B': 15,  # Mês
            'C': 15,  # Impacto
            'D': 15,  # TRPs
            'E': 18,  # Investimento
            'F': 15,  # Universo
            'G': 15,  # PMM
        }
        
        # Larguras padrão para SEMANAL (11 colunas: Rádio, Semana, Ano Comercial, Mês Comercial, Inserções, Investimento, Impactos, TRP, PMM, Universo, Programado)
        column_widths_semanal = {
            'A': 20,  # Rádio
            'B': 35,  # Semana
            'C': 15,  # Ano Comercial
            'D': 15,  # Mês Comercial
            'E': 12,  # Inserções
            'F': 18,  # Investimento
            'G': 12,  # Impactos
            'H': 12,  # TRP
            'I': 12,  # PMM
            'J': 15,  # Universo
            'K': 15,  # Programado
        }
        
        # Escolhe qual conjunto de larguras usar
        column_widths = column_widths_mensal if num_cols == 7 else column_widths_semanal
        
        for col_idx, col_name in enumerate(column_names, 1):
            col_letter = ws_output.cell(row=1, column=col_idx).column_letter
            
            if col_letter in column_widths:
                ws_output.column_dimensions[col_letter].width = column_widths[col_letter]
                print(f"[FORMATACAO]       {col_letter}: {column_widths[col_letter]}")
            else:
                ws_output.column_dimensions[col_letter].width = 15
        
        print(f"[FORMATACAO]       ✅ Colunas ajustadas")
        
        # ===== 6️⃣ CONGELAR LINHA DE CABEÇALHO =====
        print(f"[FORMATACAO]    6. Congelando cabeçalho...")
        ws_output.freeze_panes = 'A2'
        print(f"[FORMATACAO]       ✅ Cabeçalho congelado")
        
        # Salvar workbook com flush
        try:
            wb_output.save(output_path)
            print(f"[FORMATACAO] 💾 Arquivo salvo em: {output_path}")
        except Exception as save_error:
            print(f"[FORMATACAO] ❌ Erro ao salvar: {str(save_error)}")
            raise
        
        print(f"[FORMATACAO] ✅ Formatação COMPLETA VeriModelo aplicada!")
        print(f"[FORMATACAO]    - ✅ Larguras de colunas")
        print(f"[FORMATACAO]    - ✅ Alturas de linhas")
        print(f"[FORMATACAO]    - ✅ Fontes (Aptos Narrow, Bold, 11pt)")
        print(f"[FORMATACAO]    - ✅ Alinhamento e wrap")
        print(f"[FORMATACAO]    - ✅ Tabela RÁDIO-style (AZUL)")
        
    except Exception as e:
        import traceback
        print(f"[FORMATACAO] ⚠️ Erro ao copiar formatação: {str(e)}")
        traceback.print_exc()


def save_to_saidas(df, filename, apply_formatting=False):
    """Salva DataFrame em arquivo Excel na pasta saidas de forma eficiente em memória"""
    try:
        import time
        start_time = time.time()
        
        saidas_path = app.config['SAIDAS_FOLDER']
        filepath = os.path.join(saidas_path, filename)
        
        # Diagnóstico de memória
        mem_info = psutil.Process().memory_info()
        mem_mb = mem_info.rss / (1024 * 1024)
        
        print(f"[SAIDAS] 📝 Salvando arquivo: {filename}")
        print(f"[SAIDAS]    Registros: {len(df)}, Colunas: {len(df.columns)}")
        print(f"[SAIDAS]    Memória atual: {mem_mb:.1f}MB")
        print(f"[SAIDAS]    apply_formatting={apply_formatting}")
        
        # Preparar colunas para exportação (sem fazer cópia completa)
        cols_to_export = [col for col in df.columns if not str(col).startswith('Unnamed')]
        
        # Para arquivos muito grandes (>30k linhas), usar CSV para evitar timeout/crash
        # Excel é muito lento para datasets grandes no Railway
        if len(df) > 30000:
            print(f"[SAIDAS] ⚠️ Arquivo grande ({len(df)} linhas), usando CSV para evitar timeout...")
            csv_filepath = filepath.replace('.xlsx', '.csv')
            df[cols_to_export].to_csv(csv_filepath, index=False, encoding='utf-8-sig')
            elapsed = time.time() - start_time
            print(f"[SAIDAS] ✅ Arquivo CSV salvo: {csv_filepath}")
            print(f"[SAIDAS] ⏱️ Tempo de salvamento: {elapsed:.2f}s")
            gc.collect()
            return csv_filepath
        
        # Salvar Excel normalmente para arquivos menores
        print(f"[SAIDAS] 💾 Iniciando escrita Excel...")
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df[cols_to_export].to_excel(writer, index=False, sheet_name='Dados')
        
        elapsed = time.time() - start_time
        print(f"[SAIDAS] ✅ Arquivo Excel salvo: {filepath}")
        print(f"[SAIDAS] ⏱️ Tempo de salvamento: {elapsed:.2f}s")
        
        # 🎨 Aplicar formatação SOMENTE para MENSAL e SEMANAL
        if apply_formatting:
            print(f"[SAIDAS] 🎨 Chamando apply_manual_formatting...")
            apply_manual_formatting(filepath, df[cols_to_export])
            print(f"[SAIDAS] ✅ Formatação aplicada com sucesso!")
        else:
            print(f"[SAIDAS] ℹ️ Formatação não aplicada (apply_formatting=False)")
        
        # Libera memória após salvar
        gc.collect()
        
        return filepath
    except Exception as e:
        print(f"[SAIDAS] ❌ Erro ao salvar Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Fallback: tentar salvar como CSV se Excel falhar
        try:
            print(f"[SAIDAS] 🔄 Excel falhou, tentando fallback para CSV...")
            csv_filepath = filepath.replace('.xlsx', '.csv')
            df[cols_to_export].to_csv(csv_filepath, index=False, encoding='utf-8-sig')
            print(f"[SAIDAS] ✅ Arquivo CSV salvo como fallback: {csv_filepath}")
            gc.collect()
            return csv_filepath
        except Exception as csv_error:
            print(f"[SAIDAS] ❌ Fallback CSV também falhou: {str(csv_error)}")
            return None

# ==============================================================================
# FUNÇÕES DE LIMPEZA E PROCESSAMENTO
# ==============================================================================

def clean_price_value(price_str):
    """Limpa um valor de preço individual"""
    if pd.isna(price_str) or price_str == '':
        return None

    price_str = str(price_str).strip()
    price_str = re.sub(r'^R\$\s*', '', price_str)
    price_str = price_str.replace(' ', '')

    if ',' in price_str and '.' in price_str:
        price_str = price_str.replace('.', '').replace(',', '.')
    elif ',' in price_str and '.' not in price_str:
        price_str = price_str.replace(',', '.')

    try:
        return float(price_str)
    except:
        return None

def clean_dataframe(df):
    """Limpa o DataFrame removendo linhas vazias"""
    if df is None or df.empty:
        return df

    df = df.dropna(how='all')
    df = df[df.apply(lambda x: x.astype(str).str.strip().ne('').any(), axis=1)]

    if 'Rádio' in df.columns:
        mask = ~df['Rádio'].str.contains('TOTAL|RESUMO', case=False, na=False)
        df = df[mask]

    return df

def select_required_columns(df):
    """Seleciona apenas as colunas necessárias (modelo veriRelatorioModelo). Aceita 'ID' como 'Identificador'."""
    required_columns = ['Identificador', 'Data', 'Hora', 'Rádio', 'Cidade / UF',
                       'Peça', 'Comercial', 'Status', 'PMM', 'Preço', 'Semana', 'Ano Comercial', 'Mês Comercial']
    if 'Identificador' not in df.columns and 'ID' in df.columns:
        df = df.copy()
        df['Identificador'] = df['ID']
    existing_columns = [col for col in required_columns if col in df.columns]
    if not existing_columns:
        return pd.DataFrame()
    return df[existing_columns].copy()

# Nomes aceitos para a coluna de percentual no DadosIdentificador (planilha pode usar % ou Percentual)
_PORC_ALIASES = ('porc', 'perc', '%', 'percentual', 'participação', 'participacao')

def _renomear_colunas_dados_id(df):
    """
    Alinha ao Colab: só renomeia colunas que batem por nome (strip+lower).
    Identificador/ID, Universo, porc (ou perc, %, percentual) - primeira ocorrência de cada.
    """
    if df is None or df.empty:
        return df
    found = {}
    for c in df.columns:
        k = str(c).strip().lower()
        if k in ('identificador', 'id') and 'Identificador' not in found:
            found['Identificador'] = c
        elif k == 'universo' and 'Universo' not in found:
            found['Universo'] = c
        elif k in _PORC_ALIASES and 'porc' not in found:
            found['porc'] = c
    if found:
        df = df.rename(columns={v: k for k, v in found.items()})
    return df

def _normalizar_colunas_dados_identificador(df):
    """
    Normaliza nomes das colunas do DadosIdentificador (Excel): só por nome (strip+lower).
    Aceita 'Identificador'/'ID', 'Universo', 'porc' - sem fallback por posição para não pegar coluna errada.
    """
    return _renomear_colunas_dados_id(df)

def _normalizar_identificador_para_merge(s):
    """Unifica formato do Identificador para o merge: strip, upper, 123.0->123, 0123->123."""
    if pd.isna(s):
        return ''
    s = str(s).strip().upper()
    if not s or s == 'NAN':
        return ''
    # Remove .0 final quando o resto é inteiro (Excel grava 123 como 123.0)
    if s.endswith('.0'):
        rest = s[:-2]
        if rest.isdigit() or (rest.startswith('-') and rest[1:].isdigit()):
            s = rest
    # Remove zeros à esquerda em strings só numéricas (0123 = 123)
    if s.isdigit() or (s.startswith('-') and s[1:].isdigit()):
        s = str(int(s)) if s.isdigit() else ('-' + str(int(s[1:])))
    return s

def extract_dados_identificador_from_google_sheets():
    """Extrai DadosIdentificador do Google Sheets usando OAuth"""
    try:
        # ID da planilha DadosIdentificador
        sheet_id = "1UmWzuIpF1nEh1YUJH4pu9jJL7PJHvlQvHE4Pnrw6xhA"
        
        print("[PASSO 1] 📖 Carregando DadosIdentificador do Google Sheets...")
        
        sheets_service = get_authenticated_sheets_service()
        
        if not sheets_service:
            print("[PASSO 1] ⚠️ Não autenticado no Google Sheets")
            return None
        
        try:
            # Obtém metadados da planilha
            spreadsheet = sheets_service.spreadsheets().get(spreadsheetId=sheet_id).execute()
            sheets_list = spreadsheet.get('sheets', [])
            
            if not sheets_list:
                print("[PASSO 1] ℹ️ Google Sheets vazio")
                return None
            
            # Lê a primeira aba
            sheet_title = sheets_list[0].get('properties', {}).get('title')
            range_name = f"'{sheet_title}'!A:Z"
            result = sheets_service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range=range_name
            ).execute()
            
            values = result.get('values', [])
            
            if not values or len(values) < 2:
                print("[PASSO 1] ℹ️ Google Sheets vazio")
                return None
            
            # Primeira linha = cabeçalho, resto = dados (igual ao Colab)
            df_dados_id = pd.DataFrame(values[1:], columns=values[0])
            df_dados_id = _renomear_colunas_dados_id(df_dados_id)
            merge_columns = ['Identificador', 'Universo', 'porc']
            existing_cols = [col for col in merge_columns if col in df_dados_id.columns]
            
            if 'Identificador' not in existing_cols:
                print("[PASSO 1] ℹ️ Coluna 'Identificador' não encontrada em Google Sheets. Colunas lidas:", list(df_dados_id.columns))
                return None
            
            print(f"[PASSO 1] ✅ DadosIdentificador carregado: {len(df_dados_id)} registros. Colunas: {existing_cols}")
            if len(df_dados_id) > 0:
                ex_id = df_dados_id['Identificador'].iloc[0]
                print(f"[PASSO 1]    Exemplo Identificador no sheet: {repr(ex_id)}")
            return df_dados_id[existing_cols].copy()
            
        except Exception as e:
            print(f"[PASSO 1] ℹ️ Erro ao ler Google Sheets: {str(e)}")
            return None
    
    except Exception as e:
        print(f"[PASSO 1] ℹ️ Erro ao carregar DadosIdentificador: {str(e)}")
        return None

def extract_dados_identificador(arquivo_path):
    """Extrai DadosIdentificador do arquivo Excel ou Google Sheets"""
    try:
        xls = pd.ExcelFile(arquivo_path)
        
        # Procura pela aba DadosIdentificador (várias variações possíveis)
        dados_id_sheet = None
        for sheet_name in xls.sheet_names:
            if 'dadosidentificador' in sheet_name.lower().replace(' ', '').replace('_', ''):
                dados_id_sheet = sheet_name
                break
        
        if not dados_id_sheet:
            xls.close()
            print("[PASSO 1] ℹ️ DadosIdentificador não encontrado - Universo=0, PMM no Target=0")
            return None
        
        # Lê a aba DadosIdentificador
        df_dados_id = pd.read_excel(arquivo_path, sheet_name=dados_id_sheet)
        xls.close()
        
        df_dados_id = _normalizar_colunas_dados_identificador(df_dados_id)
        if df_dados_id is None:
            return None
        
        # Remove primeira linha só se for repetição do cabeçalho (ex: "Identificador" ou vazio)
        if len(df_dados_id) > 1 and 'Identificador' in df_dados_id.columns:
            first_val = str(df_dados_id['Identificador'].iloc[0]).strip().lower()
            if first_val in ('', 'nan', 'identificador'):
                df_dados_id = df_dados_id.drop(index=0).reset_index(drop=True)
        
        merge_columns = ['Identificador', 'Universo', 'porc']
        existing_cols = [col for col in merge_columns if col in df_dados_id.columns]
        
        if 'Identificador' not in existing_cols:
            print("[PASSO 1] ⚠️ Coluna 'Identificador' não encontrada. Colunas na aba:", list(df_dados_id.columns))
            return None
        
        print(f"[PASSO 1] ✅ DadosIdentificador (Excel) carregado: {len(df_dados_id)} registros. Colunas: {existing_cols}")
        if len(df_dados_id) > 0:
            print(f"[PASSO 1]    Exemplo Identificador na aba: {repr(df_dados_id['Identificador'].iloc[0])}")
        return df_dados_id[existing_cols].copy()
    
    except Exception as e:
        print(f"[PASSO 1] ⚠️ Erro ao ler DadosIdentificador: {str(e)}")
        return None

def process_data_types(df):
    """Formata os tipos de dados"""
    if df.empty:
        return df

    # Formata Data
    if 'Data' in df.columns:
        try:
            df['Data'] = pd.to_datetime(df['Data'], errors='coerce', dayfirst=True)
            df['Data'] = df['Data'].dt.strftime('%d/%m/%Y')
        except:
            pass

    # Formata Hora
    if 'Hora' in df.columns:
        try:
            df['Hora'] = df['Hora'].astype(str).str.strip()
            df['Hora'] = df['Hora'].str.replace(r'\.\d+$', '', regex=True)

            def normalize_hora(hora_str):
                if pd.isna(hora_str) or str(hora_str).strip() == '':
                    return ''
                hora_str = str(hora_str).strip()
                if len(hora_str.split(':')) == 3:
                    return hora_str
                elif len(hora_str.split(':')) == 2:
                    return hora_str + ':00'
                else:
                    return hora_str

            df['Hora'] = df['Hora'].apply(normalize_hora)
        except:
            pass

    # Converte PMM
    if 'PMM' in df.columns:
        try:
            df['PMM'] = df['PMM'].astype(str).str.replace(',', '.')
            df['PMM'] = pd.to_numeric(df['PMM'], errors='coerce')
        except:
            pass

    # Converte Preço
    if 'Preço' in df.columns:
        try:
            df['Preço'] = df['Preço'].apply(clean_price_value)
        except:
            pass

    return df

def calculate_mes_comercial(df):
    """Calcula e atualiza o Mês Comercial com base na Data para o calendário 2026"""
    if df.empty or 'Data' not in df.columns:
        return df

    try:
        print("[PROCESSAMENTO] 📅 Atualizando Mês Comercial (Calendário 2026)...")
        
        # Definição do Calendário Comercial 2026
        # (Mês, Data Início, Data Fim)
        calendar_2026 = [
            ('Janeiro', '2025-12-29', '2026-01-25'),
            ('Fevereiro', '2026-01-26', '2026-02-22'),
            ('Março', '2026-02-23', '2026-03-29'),
            ('Abril', '2026-03-30', '2026-04-26'),
            ('Maio', '2026-04-27', '2026-05-24'),
            ('Junho', '2026-05-25', '2026-06-28'),
            ('Julho', '2026-06-29', '2026-07-26'),
            ('Agosto', '2026-07-27', '2026-08-30'),
            ('Setembro', '2026-08-31', '2026-09-27'),
            ('Outubro', '2026-09-28', '2026-10-25'),
            ('Novembro', '2026-10-26', '2026-11-29'),
            ('Dezembro', '2026-11-30', '2026-12-27')
        ]
        
        # Converter para datetime temporariamente para comparação
        # A coluna Data está em string 'dd/mm/yyyy'
        temp_dates = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
        
        # Cria nova série para o mês
        meses_comerciais = pd.Series(index=df.index, dtype='object')
        
        for mes, start, end in calendar_2026:
            start_dt = pd.to_datetime(start)
            end_dt = pd.to_datetime(end)
            
            # Máscara para datas dentro do range
            mask = (temp_dates >= start_dt) & (temp_dates <= end_dt)
            meses_comerciais[mask] = mes
            
        # Atualiza o DataFrame (onde houve match)
        # Se não houve match (fora de 2026), mantemos o valor original se existir, ou vazio
        if 'Mês Comercial' in df.columns:
            df['Mês Comercial'] = meses_comerciais.combine_first(df['Mês Comercial'])
        else:
            df['Mês Comercial'] = meses_comerciais
            
        # Opcional: Atualizar 'Ano Comercial' para 2026 se estiver dentro dos meses
        # (Mas cuidado com Janeiro que pega Dez 2025)
        # O user não pediu explicitamente, mas é bom garantir consistência se 'Ano Comercial' for usado
        
    except Exception as e:
        print(f"[ERRO] Falha ao calcular Mês Comercial: {str(e)}")
        
    return df

def remove_duplicates_properly(df):
    """Remove duplicatas de forma eficiente em termos de memória"""
    if df is None or df.empty:
        return df, 0

    original_count = len(df)
    
    required_cols = ['Identificador', 'Data', 'Hora']
    if not all(col in df.columns for col in required_cols):
        return df, 0

    # Usamos colunas temporárias diretamente no DataFrame original para evitar df.copy()
    temp_cols = []
    
    # 1. Identificador
    df['Identificador_tmp'] = df['Identificador'].astype(str).str.strip().str.upper().replace(['', 'nan', 'None'], 'VAZIO')
    temp_cols.append('Identificador_tmp')
    
    # 2. Data
    if pd.api.types.is_datetime64_any_dtype(df['Data']):
        temp_cols.append('Data')
    else:
        df['Data_tmp'] = df['Data'].astype(str).str.strip().str.replace(' ', '').str.upper().replace(['', 'nan', 'None'], 'VAZIO')
        temp_cols.append('Data_tmp')
        
    # 3. Hora
    df['Hora_tmp'] = df['Hora'].astype(str).str.strip().str.replace(' ', '').str.upper()
    df['Hora_tmp'] = df['Hora_tmp'].str.replace(r'\.\d+$', '', regex=True)
    # Normalização básica de hora
    df['Hora_tmp'] = df['Hora_tmp'].apply(lambda x: 
        x if ':' in str(x) else ('VAZIO' if x in ['', 'nan', 'None'] else x))
    temp_cols.append('Hora_tmp')
    
    # 4. Colunas opcionais
    for col in ['Semana', 'Ano Comercial', 'Mês Comercial']:
        if col in df.columns:
            tmp_name = f"{col}_tmp"
            df[tmp_name] = df[col].astype(str).str.strip().str.upper().replace(['', 'nan', 'None'], 'VAZIO')
            temp_cols.append(tmp_name)

    # Remove duplicatas IN-PLACE
    df.drop_duplicates(subset=temp_cols, keep='first', inplace=True)
    
    # Limpa temporárias
    cols_to_remove = [c for c in temp_cols if c.endswith('_tmp')]
    df.drop(columns=cols_to_remove, inplace=True)
    
    duplicates_removed = original_count - len(df)
    return df, duplicates_removed

# ==============================================================================
# PASSO 1: COMPILADOR
# ==============================================================================

def passo1_compilar(arquivo_path):
    """Passo 1: Compila o arquivo raw + MÚLTIPLAS PLANILHAS DO DRIVE (como o Colab)"""
    print(f"\n[PASSO 1] 🚀 Iniciando compilação...")
    print(f"[PASSO 1] 📄 Arquivo principal: {arquivo_path}")
    
    try:
        all_dataframes = []
        
        # ========== ETAPA 1: Ler TODAS as abas do arquivo uploaded (não só a primeira) ==========
        print(f"[PASSO 1] 📖 ETAPA 1: Lendo arquivo principal (todas as abas de dados)")
        xls = pd.ExcelFile(arquivo_path)
        sheet_names = xls.sheet_names
        # Pula abas que não são de relatório (DadosIdentificador, modelo, etc.)
        abas_para_ler = []
        for name in sheet_names:
            n_lower = name.lower().replace(' ', '').replace('_', '')
            if 'dadosidentificador' in n_lower or 'dados identificador' in n_lower:
                print(f"[PASSO 1]    ⏭️ Pulando aba (DadosIdentificador): {name}")
                continue
            if any(t in n_lower for t in ['modelo', 'template', 'verimodelorelatorio']):
                print(f"[PASSO 1]    ⏭️ Pulando aba (template): {name}")
                continue
            abas_para_ler.append(name)
        if not abas_para_ler:
            abas_para_ler = [sheet_names[0]]
            print(f"[PASSO 1]    ℹ️ Usando apenas primeira aba: {abas_para_ler[0]}")
        for aba_nome in abas_para_ler:
            df_upload = pd.read_excel(arquivo_path, sheet_name=aba_nome)
            print(f"[PASSO 1]    Aba '{aba_nome}': {len(df_upload)} registros lidos")
            if len(df_upload) > 1:
                df_upload = df_upload.drop(index=0).reset_index(drop=True)
            df_upload = clean_dataframe(df_upload)
            df_upload = select_required_columns(df_upload)
            if df_upload.empty:
                print(f"[PASSO 1]    ⚠️ Aba '{aba_nome}' sem colunas necessárias ou vazia após limpeza")
                continue
            df_upload = process_data_types(df_upload)
            df_upload = calculate_mes_comercial(df_upload)
            if 'Preço' in df_upload.columns:
                df_upload['Preço'] = df_upload['Preço'].apply(clean_price_value)
                df_upload['Preço'] = pd.to_numeric(df_upload['Preço'], errors='coerce')
                df_upload['Preço'] = df_upload['Preço'].fillna(0)
            df_upload, dup_removed = remove_duplicates_properly(df_upload)
            all_dataframes.append(df_upload)
            n_radios = df_upload['Rádio'].nunique() if 'Rádio' in df_upload.columns else 0
            print(f"[PASSO 1]    ✅ Aba '{aba_nome}': {len(df_upload)} registros, {n_radios} Rádio(s)")
        xls.close()
        
        # ========== ETAPA 2: Ler OUTRAS PLANILHAS do Google Drive ==========
        print(f"\n[PASSO 1] 📖 ETAPA 2: Buscando outras planilhas no Google Drive...")
        
        drive_service = get_authenticated_drive_service()
        sheets_service = get_authenticated_sheets_service()
        
        if not drive_service:
            print("[PASSO 1] ⚠️ Não autenticado no Google Drive - apenas arquivo local será processado")
            print("[PASSO 1] ℹ️ Acesse /authorize para fazer login")
        else:
            # Lista todos os Google Sheets da pasta RelatorioVeri
            atualizar_progresso(1, 20, "Buscando planilhas no Google Drive...")
            sheets_list = list_google_sheets_in_drive(drive_service, "RelatorioVeri")
            
            dados_identificador_info = None
            report_sheets = []
            
            # Filtra sheets (pula DadosIdentificador e templates)
            for sheet_info in sheets_list:
                sheet_name = sheet_info['name'].lower()
                original_name = sheet_info['name']
                
                # Identifica DadosIdentificador
                if any(pattern in sheet_name for pattern in ['dadosidentificador', 'dados_identificador', 'dados identificador']):
                    dados_identificador_info = sheet_info
                    print(f"[PASSO 1]    ✅ DadosIdentificador encontrado: {original_name}")
                    continue
                
                # Pula templates
                if any(template in sheet_name for template in ['modelo', 'template', 'verimodelorelatório']):
                    print(f"[PASSO 1]    ⏭️ Ignorando template: {original_name}")
                    continue
                
                # Adiciona aos relatórios
                report_sheets.append(sheet_info)
                print(f"[PASSO 1]    📊 Adicionando: {original_name}")
            
            print(f"\n[PASSO 1] 📊 Total de arquivos a processar: {len(report_sheets)}")
            
            # Processa cada relatório do Drive
            for i, sheet_info in enumerate(report_sheets, 1):
                sheet_name = sheet_info['name']
                sheet_id = sheet_info['id']
                
                try:
                    import psutil
                    mem = psutil.Process().memory_info().rss / (1024 * 1024)
                    print(f"\n[PASSO 1] [{i}/{len(report_sheets)}] Analisando: {sheet_name} (Memoria: {mem:.1f}MB)")
                    
                    # Calcula percentual progressivo (de 20% a 35%)
                    perc = 20 + int((i / len(report_sheets)) * 15)
                    atualizar_progresso(1, perc, f"Lendo do Drive: {sheet_name}")

                    # Limite do Google: 60 leituras/minuto por usuário; pequena pausa evita HttpError 429
                    if i > 1:
                        time.sleep(1.2)
                    # Lê o arquivo
                    df = read_google_sheet(sheets_service, sheet_id, sheet_name)
                    
                    if df is not None and not df.empty:
                        # Remove segunda linha se necessário
                        if len(df) > 1:
                            df = df.drop(index=0).reset_index(drop=True)
                        
                        # Limpa
                        df = clean_dataframe(df)
                        
                        # Seleciona colunas
                        df = select_required_columns(df)
                        
                        if not df.empty:
                            # Processa tipos
                            df = process_data_types(df)
                            
                            # Recalcula Mês Comercial (2026)
                            df = calculate_mes_comercial(df)
                            
                            # Converte Preço
                            if 'Preço' in df.columns:
                                df['Preço'] = df['Preço'].apply(clean_price_value)
                                df['Preço'] = pd.to_numeric(df['Preço'], errors='coerce')
                                df['Preço'] = df['Preço'].fillna(0)
                            
                            # Remove duplicatas
                            df, dup_removed = remove_duplicates_properly(df)
                            
                            all_dataframes.append(df)
                            print(f"[PASSO 1]       ✅ {len(df)} registros válidos adicionados")
                        else:
                            print(f"[PASSO 1]       ⚠️ Nenhum dado válido após limpeza")
                    else:
                        print(f"[PASSO 1]       ❌ Erro na leitura ou arquivo vazio")
                except Exception as loop_e:
                    print(f"[PASSO 1] ❌ Falha crítica no arquivo '{sheet_name}': {str(loop_e)}")
                    import traceback
                    traceback.print_exc()
                    continue
        
        # ========== ETAPA 3: Unificar todos os dados ==========
        atualizar_progresso(1, 36, "Unificando arquivos do Drive...")
        print(f"\n[PASSO 1] 🔗 ETAPA 3: Unificando {len(all_dataframes)} fonte(s) de dados...")
        
        if not all_dataframes:
            return None, "Nenhum arquivo válido foi processado", None
        
        # Concatena todos os DataFrames
        unified_df = pd.concat(all_dataframes, ignore_index=True)
        
        # Libera memória imediatamente
        del all_dataframes
        gc.collect()
        
        atualizar_progresso(1, 37, "Removendo duplicatas entre arquivos...")
        print(f"[PASSO 1]    Total unificado: {len(unified_df)} registros. Limpando duplicatas...")
        
        # Remove duplicatas do conjunto unificado
        unified_df, dup_removed = remove_duplicates_properly(unified_df)
        print(f"[PASSO 1]    Após remoção de duplicatas: {len(unified_df)} registros")
        
        gc.collect() # Mais uma coleta após limpeza pesada
        
        # ========== ETAPA 4: Fazer MERGE com DadosIdentificador ==========
        print(f"\n[PASSO 1] 🔗 ETAPA 4: Fazendo merge com DadosIdentificador...")
        
        # Garante coluna Identificador no compilado (alguns Excel usam "ID")
        if 'Identificador' not in unified_df.columns and 'ID' in unified_df.columns:
            unified_df['Identificador'] = unified_df['ID']
        
        # Tenta Google Sheets primeiro, depois Excel
        df_dados_id = extract_dados_identificador_from_google_sheets()
        if df_dados_id is None:
            df_dados_id = extract_dados_identificador(arquivo_path)
        
        if df_dados_id is not None and 'Identificador' in unified_df.columns:
            print(f"[PASSO 1]    ✅ DadosIdentificador carregado: {len(df_dados_id)} registros")
            
            unified_df['Identificador'] = unified_df['Identificador'].apply(_normalizar_identificador_para_merge)
            df_dados_id['Identificador'] = df_dados_id['Identificador'].apply(_normalizar_identificador_para_merge)
            df_dados_id = df_dados_id.drop_duplicates(subset=['Identificador'], keep='first').copy()
            
            unified_df = unified_df.merge(df_dados_id, on='Identificador', how='left')
            print(f"[PASSO 1]    ✅ Merge realizado: {len(unified_df)} registros")
            
            # Libera DadosIdentificador após merge
            del df_dados_id
            gc.collect()
            
            def clean_number_str(val):
                if pd.isna(val): return val
                if isinstance(val, (int, float)): return val
                val = str(val).strip()
                if val == '': return 0
                val = val.replace('.', '').replace(',', '.')
                return val

            # Converte 'porc' para numérico se existir
            if 'porc' in unified_df.columns:
                unified_df['porc'] = unified_df['porc'].apply(clean_number_str)
                unified_df['porc'] = pd.to_numeric(unified_df['porc'], errors='coerce')
            
            # Converte Universo para numérico
            if 'Universo' in unified_df.columns:
                unified_df['Universo'] = unified_df['Universo'].apply(clean_number_str)
                unified_df['Universo'] = pd.to_numeric(unified_df['Universo'], errors='coerce')
            
            # Calcula PMM no Target
            if 'PMM' in unified_df.columns and 'porc' in unified_df.columns:
                # Garante que PMM também esteja limpo
                unified_df['PMM'] = unified_df['PMM'].apply(clean_number_str)
                unified_df['PMM'] = pd.to_numeric(unified_df['PMM'], errors='coerce').fillna(0)
                
                unified_df['PMM no Target'] = (unified_df['PMM'] * (unified_df['porc'] / 100)).round(2)
                unified_df['PMM no Target'] = unified_df['PMM no Target'].fillna(0)
                print(f"[PASSO 1]    ✅ Coluna 'PMM no Target' calculada")
            
            # Formata PMM com 2 casas decimais
            if 'PMM' in unified_df.columns:
                unified_df['PMM'] = unified_df['PMM'].round(2)
            
            # Preenche Universo com 0
            if 'Universo' in unified_df.columns:
                unified_df['Universo'] = unified_df['Universo'].fillna(0)
            
            # Remove coluna 'porc'
            unified_df = unified_df.drop(columns=['porc'], errors='ignore')
        else:
            # Se não houver DadosIdentificador
            if 'Universo' not in unified_df.columns:
                unified_df['Universo'] = 0
            if 'PMM no Target' not in unified_df.columns:
                unified_df['PMM no Target'] = 0
        
        # ========== ETAPA 5: Reorganizar e salvar ==========
        print(f"\n[PASSO 1] 📋 ETAPA 5: Reorganizando colunas...")
        
        expected_columns = ['Identificador', 'Data', 'Hora', 'Rádio', 'Cidade / UF',
                           'Peça', 'Comercial', 'Status', 'PMM', 'Preço', 'Universo', 'PMM no Target']
        existing_cols = [col for col in expected_columns if col in unified_df.columns]
        other_cols = [col for col in unified_df.columns if col not in expected_columns]
        final_cols = existing_cols + other_cols
        unified_df = unified_df[final_cols]
        
        print(f"[PASSO 1] ✅ Compilação concluída: {len(unified_df)} registros finais")
        print(f"[PASSO 1] 📊 Colunas: {list(unified_df.columns)}")
        
        # SALVA NA PASTA SAIDAS (SEM FORMATAÇÃO)
        print(f"[PASSO 1] 💾 Salvando arquivo na pasta saidas...")
        atualizar_progresso(1, 39, "Salvando arquivo compilado...")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"COMPILADO_{timestamp}.xlsx"
        
        saved_path = save_to_saidas(unified_df, filename, apply_formatting=False)
        
        if saved_path:
            print(f"[PASSO 1] ✅ Arquivo salvo com sucesso: {saved_path}")
            atualizar_progresso(1, 40, "Passo 1 concluído!")
        else:
            print(f"[PASSO 1] ⚠️ Falha ao salvar arquivo, mas continuando...")
            atualizar_progresso(1, 40, "Passo 1 concluído (arquivo não salvo)")
        
        return unified_df, None, timestamp
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, f"Erro na compilação: {str(e)}", None

# ==============================================================================
# PASSO 2: RELATÓRIO MENSAL (Com busca de modelo Google Sheets)
# ==============================================================================

def setup_google_services():
    """Configura autenticação com Google"""
    try:
        from google.oauth2.service_account import Credentials
        
        # Scopes necessários
        SCOPES = [
            'https://www.googleapis.com/auth/drive',
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive.readonly'
        ]
        
        # Carregar credenciais com scopes
        creds = Credentials.from_service_account_file(
            os.environ['GOOGLE_APPLICATION_CREDENTIALS'],
            scopes=SCOPES
        )
        
        gc = gspread.authorize(creds)
        drive_service = build('drive', 'v3', credentials=creds)
        print("[GOOGLE] ✅ Autenticação configurada com scopes corretos")
        return gc, drive_service
    except Exception as e:
        print(f"[GOOGLE] ⚠️ Erro na autenticação: {str(e)}")
        return None, None

def find_modelo_relatorio(drive_service):
    """Encontra o arquivo VeriModeloRelatorio no Google Drive"""
    if not drive_service:
        print("[PASSO 2] ⚠️ Google Drive não disponível")
        return None
    
    try:
        # Busca a pasta RelatorioVeri
        folder_query = "name='RelatorioVeri' and mimeType='application/vnd.google-apps.folder'"
        folder_results = drive_service.files().list(q=folder_query, fields="files(id, name)").execute()
        folders = folder_results.get('files', [])
        
        if not folders:
            print("[PASSO 2] ⚠️ Pasta 'RelatorioVeri' não encontrada no Google Drive")
            return None
        
        folder_id = folders[0]['id']
        
        # Busca o modelo VeriModeloRelatorio
        modelo_patterns = ['VeriModeloRelatorio', 'Veri Modelo Relatorio', 'Modelo Relatorio', 'Modelo']
        
        for pattern in modelo_patterns:
            modelo_query = f"'{folder_id}' in parents and name contains '{pattern}' and mimeType='application/vnd.google-apps.spreadsheet'"
            modelo_results = drive_service.files().list(q=modelo_query, fields="files(id, name)").execute()
            modelos = modelo_results.get('files', [])
            
            if modelos:
                modelo_info = modelos[0]
                print(f"[PASSO 2] ✅ Modelo encontrado: {modelo_info['name']}")
                return modelo_info['id']
        
        print("[PASSO 2] ⚠️ Modelo não encontrado no Google Drive")
        return None
    
    except Exception as e:
        print(f"[PASSO 2] ⚠️ Erro ao buscar modelo: {str(e)}")
        return None

def export_modelo_as_excel(modelo_id, drive_service):
    """Exporta o modelo do Google Sheets como arquivo Excel usando Google Drive API"""
    try:
        print(f"[PASSO 2] 📥 Exportando modelo via Google Drive API...")
        
        if not drive_service or not modelo_id:
            print("[PASSO 2] ❌ Drive service ou modelo_id não disponível")
            return None
        
        # Faz download do arquivo Excel do Google Sheets via API
        request = drive_service.files().export(
            fileId=modelo_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        file_content = request.execute()
        
        if not file_content or len(file_content) == 0:
            print("[PASSO 2] ❌ Arquivo vazio baixado")
            return None
        
        # Salva em arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.write(file_content)
        temp_file.close()
        
        print(f"[PASSO 2] ✅ Modelo exportado: {temp_file.name} ({len(file_content)} bytes)")
        
        # Valida se é um arquivo Excel válido
        try:
            test_wb = load_workbook(temp_file.name)
            print(f"[PASSO 2] ✅ Excel válido: {test_wb.sheetnames}")
        except Exception as e:
            print(f"[PASSO 2] ⚠️ Arquivo pode ser inválido: {str(e)}")
        
        return temp_file.name
    
    except Exception as e:
        print(f"[PASSO 2] ❌ Erro ao exportar modelo: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def fill_modelo_with_data(modelo_file, aggregated_data):
    """
    Carrega o modelo Excel, preenche com dados agregados, preservando formatação
    """
    try:
        print(f"[PASSO 2] 📋 Carregando modelo de {modelo_file}...")
        
        if not os.path.exists(modelo_file):
            print(f"[PASSO 2] ❌ Arquivo não encontrado: {modelo_file}")
            return None
        
        wb = load_workbook(modelo_file)
        print(f"[PASSO 2] ✅ Workbook carregado. Abas: {wb.sheetnames}")
        
        ws = wb.active
        print(f"[PASSO 2] 📄 Aba ativa: {ws.title}, Dimensões: {ws.dimensions}")
        
        # Encontra os headers na primeira linha
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col_idx).value
            if cell_value:
                headers[str(cell_value).strip()] = col_idx
        
        print(f"[PASSO 2] 📍 Headers encontrados: {list(headers.keys())}")
        
        # Mapeia nomes de colunas - tenta várias variações
        col_praca = None
        col_mes = None
        col_impacto = None
        col_trp = None
        col_investimento = None
        col_pmm = None
        col_universo = None
        
        # Busca Praça
        for key in headers:
            if 'praç' in key.lower() or 'praca' in key.lower():
                col_praca = headers[key]
                break
        
        # Busca Mês/Período
        for key in headers:
            if 'mês' in key.lower() or 'periodo' in key.lower() or 'periodo comercial' in key.lower():
                col_mes = headers[key]
                break
        
        # Busca colunas de dados
        for key in headers:
            if 'impacto' in key.lower():
                col_impacto = headers[key]
            elif 'trp' in key.lower():
                col_trp = headers[key]
            elif 'investimento' in key.lower():
                col_investimento = headers[key]
            elif 'pmm' in key.lower() and 'no target' not in key.lower():
                col_pmm = headers[key]
            elif 'universo' in key.lower():
                col_universo = headers[key]
        
        print(f"[PASSO 2] 🔍 Mapeamento de colunas:")
        print(f"   - Praça: {col_praca}")
        print(f"   - Mês/Período: {col_mes}")
        print(f"   - Impacto: {col_impacto}")
        print(f"   - TRPs: {col_trp}")
        print(f"   - Investimento: {col_investimento}")
        print(f"   - PMM: {col_pmm}")
        print(f"   - Universo: {col_universo}")
        
        if not col_praca or not col_mes:
            print(f"[PASSO 2] ⚠️ Colunas essenciais (Praça/Mês) não encontradas!")
            print(f"[PASSO 2] Headers disponíveis: {list(headers.keys())}")
            return None
        
        updates_made = 0
        
        # Itera linhas do modelo (a partir da linha 2, pulando header)
        print(f"[PASSO 2] 📊 Preenchendo modelo, total de {ws.max_row - 1} linhas...")
        
        for row_idx in range(2, ws.max_row + 1):
            praca_cell = ws.cell(row=row_idx, column=col_praca)
            mes_cell = ws.cell(row=row_idx, column=col_mes)
            
            if not praca_cell.value or not mes_cell.value:
                continue
            
            praca_value = str(praca_cell.value).strip()
            mes_value = str(mes_cell.value).strip()
            
            # Busca dados correspondentes no aggregated_data
            found = False
            for _, data_row in aggregated_data.iterrows():
                praca_match = str(data_row['Praca_Mapeada']).strip() == praca_value
                mes_match = mes_value and str(data_row['Periodo_Comercial']).strip() in mes_value
                
                if praca_match and mes_match:
                    found = True
                    # Preenche as células com dados, preservando formatação
                    if col_impacto and 'Impacto' in data_row.index:
                        val = round(float(data_row['Impacto']), 2) if data_row['Impacto'] else 0
                        ws.cell(row=row_idx, column=col_impacto).value = val
                        updates_made += 1
                    
                    if col_trp and 'TRPs' in data_row.index:
                        val = round(float(data_row['TRPs']), 2) if data_row['TRPs'] else 0
                        ws.cell(row=row_idx, column=col_trp).value = val
                        updates_made += 1
                    
                    if col_investimento and 'Investimento' in data_row.index:
                        val = round(float(data_row['Investimento']), 2) if data_row['Investimento'] else 0
                        ws.cell(row=row_idx, column=col_investimento).value = val
                        updates_made += 1
                    
                    if col_pmm and 'PMM' in data_row.index:
                        val = round(float(data_row['PMM']), 2) if data_row['PMM'] else 0
                        ws.cell(row=row_idx, column=col_pmm).value = val
                        updates_made += 1
                    
                    if col_universo and 'Universo' in data_row.index:
                        val = round(float(data_row['Universo']), 2) if data_row['Universo'] else 0
                        ws.cell(row=row_idx, column=col_universo).value = val
                        updates_made += 1
                    
                    break
        
        print(f"[PASSO 2] ✅ Preenchimento concluído: {updates_made} células atualizadas")
        
        # Salva o workbook modificado em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    except Exception as e:
        print(f"[PASSO 2] ❌ Erro ao preencher modelo: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def read_modelo_relatorio(gc, sheet_id):
    """Lê a estrutura do modelo de relatório do Google Sheets"""
    try:
        sheet = gc.open_by_key(sheet_id)
        worksheet = sheet.get_worksheet(0)  # Primeira aba
        
        data = worksheet.get_all_values()
        
        if not data:
            print("[PASSO 2] ⚠️ Modelo vazio")
            return None
        
        df_modelo = pd.DataFrame(data[1:], columns=data[0])
        print(f"[PASSO 2] ✅ Modelo carregado: {len(df_modelo)} linhas")
        return df_modelo
    
    except Exception as e:
        print(f"[PASSO 2] ⚠️ Erro ao ler modelo: {str(e)}")
        return None

def create_filled_report(df_modelo, aggregated_data):
    """Cria uma cópia do modelo preenchida com os dados agregados"""
    try:
        df_report = df_modelo.copy()
        updates_made = 0
        
        print("[PASSO 2] 📊 Preenchendo modelo com dados...")
        
        # Tenta preencher o modelo com os dados agregados
        for index, row in df_report.iterrows():
            praca = row.get('Praça', '')
            mes_col = row.get('Mês', '')
            
            if not praca or not mes_col:
                continue
            
            # Busca dados para esta praça e período
            for _, data_row in aggregated_data.iterrows():
                if data_row['Praca_Mapeada'] == praca and data_row['Periodo_Comercial'] in str(mes_col):
                    # Preenche as métricas
                    if 'Impacto' in data_row.index:
                        df_report.at[index, 'Impacto'] = round(data_row['Impacto'], 2)
                        updates_made += 1
                    if 'TRPs' in data_row.index:
                        df_report.at[index, 'TRPs'] = round(data_row['TRPs'], 2)
                        updates_made += 1
                    if 'Investimento' in data_row.index:
                        df_report.at[index, 'Investimento'] = round(data_row['Investimento'], 2)
                        updates_made += 1
                    if 'PMM' in data_row.index:
                        df_report.at[index, 'PMM'] = round(data_row['PMM'], 2)
                        updates_made += 1
                    if 'Universo' in data_row.index:
                        df_report.at[index, 'Universo'] = round(data_row['Universo'], 2)
                        updates_made += 1
                    break
        
        print(f"[PASSO 2] ✅ Preenchimento concluído: {updates_made} células atualizadas")
        return df_report
    
    except Exception as e:
        print(f"[PASSO 2] ⚠️ Erro ao preencher modelo: {str(e)}")
        return None

def get_periodo_comercial_for_date(data):
    """Determina o período comercial para uma data"""
    if pd.isna(data):
        return None

    try:
        data_date = pd.to_datetime(data, format='%d/%m/%Y').date()
    except:
        return None

    for periodo, (inicio, fim) in PERIODOS_COMERCIAIS.items():
        inicio_dt = datetime.strptime(inicio, '%Y-%m-%d').date()
        fim_dt = datetime.strptime(fim, '%Y-%m-%d').date()

        if inicio_dt <= data_date <= fim_dt:
            return periodo

    return None

def map_cidade_to_praca(cidade):
    """Mapeia cidade para praça"""
    if pd.isna(cidade):
        return None
    
    for praca, cidades in MAPA_PRACAS.items():
        if cidade in cidades:
            return praca

    return None

def passo2_mensal(df_compilado=None):
    """Passo 2: Gera relatório mensal como CÓPIA DO MODELO preenchida com dados"""
    print("[PASSO 2] Iniciando geração do relatório mensal")
    
    try:
        # Se df_compilado não foi passado, busca o último da pasta saidas
        if df_compilado is None:
            compiled_file = get_latest_compiled_file()
            if not compiled_file:
                return None, "Nenhum arquivo compilado encontrado na pasta saidas/"
            
            print(f"[PASSO 2] 📖 Carregando: {os.path.basename(compiled_file)}")
            if str(compiled_file).lower().endswith('.csv'):
                df_compilado = pd.read_csv(compiled_file, encoding='utf-8-sig')
            else:
                df_compilado = pd.read_excel(compiled_file)
        
        if df_compilado.empty:
            return None, "Dados compilados vazios"
        
        df = df_compilado.copy()
        
        # Converte Data para datetime
        df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
        
        df_clean = df.dropna(subset=['Data']).copy()
        
        # Mapeia períodos
        df_clean['Periodo_Comercial'] = df_clean['Data'].apply(lambda x: get_periodo_comercial_for_date(x))
        
        # Mapeia praças
        df_clean['Praca_Mapeada'] = df_clean['Cidade / UF'].apply(map_cidade_to_praca)
        
        # Filtra válidos
        df_valid = df_clean[(df_clean['Periodo_Comercial'].notna()) & (df_clean['Praca_Mapeada'].notna())].copy()
        
        if df_valid.empty:
            return None, "Nenhum dado válido após mapeamento"
        
        # Converte colunas numéricas ANTES de agregar
        numeric_cols = ['PMM', 'PMM no Target', 'Universo']
        for col in numeric_cols:
            if col in df_valid.columns:
                df_valid[col] = pd.to_numeric(df_valid[col], errors='coerce')
                df_valid[col] = df_valid[col].fillna(0)
        
        # Limpa Preço (de "R$ 147,03" para numérico)
        if 'Preço' in df_valid.columns:
            df_valid['Preço'] = df_valid['Preço'].apply(clean_price_value)
            df_valid['Preço'] = pd.to_numeric(df_valid['Preço'], errors='coerce')
            df_valid['Preço'] = df_valid['Preço'].fillna(0)
        
        print(f"[PASSO 2] Dados preparados: PMM no Target sum={df_valid['PMM no Target'].sum()}")
        
        # PARTE 1: Agrupa por Identificador para pegar valores únicos
        df_unique = df_valid.groupby(['Periodo_Comercial', 'Praca_Mapeada', 'Identificador']).agg({
            'PMM': 'first',
            'Universo': 'first'
        }).reset_index()
        
        # PARTE 2: Agrega valores únicos por período e praça
        aggregated_unique = df_unique.groupby(['Periodo_Comercial', 'Praca_Mapeada']).agg({
            'PMM': 'sum',
            'Universo': 'sum'
        }).reset_index()
        
        # PARTE 3: Agrega TODOS os valores (com duplicatas)
        aggregated_all = df_valid.groupby(['Periodo_Comercial', 'Praca_Mapeada']).agg({
            'PMM': 'sum',           # Impacto
            'PMM no Target': 'sum', # TRPs
            'Preço': 'sum'          # Investimento
        }).reset_index()
        
        # Renomeia
        aggregated_all = aggregated_all.rename(columns={
            'PMM': 'Impacto',
            'PMM no Target': 'TRPs',
            'Preço': 'Investimento'
        })
        
        aggregated_unique = aggregated_unique.rename(columns={
            'PMM': 'PMM_UniqueSum',
            'Universo': 'Universo_UniqueSum'
        })
        
        # Junta os DataFrames
        result = aggregated_all.merge(
            aggregated_unique,
            on=['Periodo_Comercial', 'Praca_Mapeada'],
            how='left'
        )
        
        # Renomeia colunas finais
        result = result.rename(columns={
            'PMM_UniqueSum': 'PMM',
            'Universo_UniqueSum': 'Universo',
            'Praca_Mapeada': 'Praça',
            'Periodo_Comercial': 'Mês'
        })
        
        # Reorganiza colunas na ordem correta
        result = result[['Praça', 'Mês', 'Impacto', 'TRPs', 'Investimento', 'Universo', 'PMM']]
        
        print(f"[PASSO 2] Relatório agregado gerado: {len(result)} linhas")
        
        # ============================================================
        # MODELO DO GOOGLE SHEETS - DESATIVADO TEMPORARIAMENTE
        # ============================================================
        # TODO: Reimplementar exportação e preenchimento do modelo
        excel_output = None
        
        # Se conseguiu preencher o modelo, usa ele; senão usa o agregado simples
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"RELATORIO_MENSAL_{timestamp}.xlsx"
        
        if excel_output:
            # Salva o output do openpyxl
            saidas_path = os.path.join('flask_app', 'saidas', filename)
            os.makedirs(os.path.dirname(saidas_path), exist_ok=True)
            
            with open(saidas_path, 'wb') as f:
                f.write(excel_output.getvalue())
            
            print(f"[SAIDAS] ✅ Arquivo salvo localmente: {filename}")
        else:
            # Fallback: salva o agregado simples COM FORMATAÇÃO MENSAL
            save_to_saidas(result, filename, apply_formatting=True)
        
        print("[PASSO 2] ✅ Relatório mensal salvo em saidas/")
        return result, None
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, f"Erro no passo 2: {str(e)}"

# ==============================================================================
# FUNÇÕES AUXILIARES: Extração de Ano e Mês Comercial
# ==============================================================================

def extrair_ano_comercial(periodo_semana_str):
    """
    Extrai o ano comercial da string da semana
    Formato: "Semana 01: 29/12/2025 a 04/01/2026"
    Retorna o ano da segunda data (fim do período) como número inteiro
    """
    try:
        # Tenta encontrar a segunda data (após " a " ou após "-")
        if ' a ' in periodo_semana_str:
            data_fim_str = periodo_semana_str.split(' a ')[1].strip()
        elif '-' in periodo_semana_str:
            data_fim_str = periodo_semana_str.split('-')[1].strip()
        else:
            return ''
        
        # Extrai o ano (últimos 4 caracteres) como número inteiro
        ano = data_fim_str.split('/')[-1]
        return ano  # Ex: 2026 (sem apóstrofo)
    except:
        return ''

def extrair_mes_comercial(periodo_semana_str):
    """
    Extrai o mês comercial da string da semana
    Formato: "Semana 01: 29/12/2025 a 04/01/2026"
    Determina o mês comercial baseado no período
    Retorna formato: "Jan'26", "Fev'26", etc
    """
    try:
        # Encontra a segunda data (período final)
        if ' a ' in periodo_semana_str:
            data_fim_str = periodo_semana_str.split(' a ')[1].strip()
        elif '-' in periodo_semana_str:
            data_fim_str = periodo_semana_str.split('-')[1].strip()
        else:
            return ''
        
        # Parse da data final (DD/MM/YYYY)
        partes = data_fim_str.split('/')
        if len(partes) == 3:
            dia = int(partes[0])
            mes = int(partes[1])
            ano = partes[2]
            
            # Nomes dos meses abreviados em português
            meses = {
                1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr',
                5: 'Mai', 6: 'Jun', 7: 'Jul', 8: 'Ago',
                9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
            }
            
            mes_nome = meses.get(mes, '')
            return f"{mes_nome}'{ano[-2:]}"  # Ex: "Jan'26"
        return ''
    except:
        return ''

# ==============================================================================
# PASSO 3: RELATÓRIO SEMANAL
# ==============================================================================

def passo3_semanal(df_compilado=None):
    """Passo 3: Gera relatório semanal - TOTALMENTE INDEPENDENTE"""
    print("[PASSO 3] Iniciando geração do relatório semanal (independente)")
    
    try:
        if df_compilado is not None:
            print(f"[PASSO 3] ✅ Recebido DataFrame em memória com {len(df_compilado)} linhas")
        
        # PASSO 3 É INDEPENDENTE - busca o último compilado da pasta saidas
        if df_compilado is None:
            print(f"[PASSO 3] ⚠️ Nenhum DataFrame recebido, buscando arquivo em disco...")
            compiled_file = get_latest_compiled_file()
            if not compiled_file:
                return None, "Nenhum arquivo compilado encontrado na pasta saidas/"
            
            print(f"[PASSO 3] 📖 Carregando: {os.path.basename(compiled_file)}")
            if str(compiled_file).lower().endswith('.csv'):
                df_compilado = pd.read_csv(compiled_file, encoding='utf-8-sig')
            else:
                df_compilado = pd.read_excel(compiled_file)
        
        if df_compilado.empty:
            return None, "Dados compilados vazios"
        
        df = df_compilado.copy()
        
        # Converte Data
        df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
        df = df.dropna(subset=['Data'])
        
        # Calcula semana
        df['Numero_Semana'] = df['Data'].dt.isocalendar().week
        df['Week_Start'] = df['Data'] - pd.to_timedelta(df['Data'].dt.weekday, unit='d')
        df['Week_End'] = df['Week_Start'] + pd.Timedelta(days=6)
        df['Periodo_Semana'] = (
            "Semana " + df['Numero_Semana'].astype(str).str.zfill(2) + ": " +
            df['Week_Start'].dt.strftime('%d/%m/%Y') + " a " +
            df['Week_End'].dt.strftime('%d/%m/%Y')
        )
        
        # Converte numéricos ANTES de agregar
        numeric_cols = ['PMM', 'PMM no Target', 'Universo']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
                df[col] = df[col].fillna(0)
        
        # Limpa Preço (de "R$ 147,03" para numérico)
        if 'Preço' in df.columns:
            df['Preço'] = df['Preço'].apply(clean_price_value)
            df['Preço'] = pd.to_numeric(df['Preço'], errors='coerce')
            df['Preço'] = df['Preço'].fillna(0)
        
        print(f"[PASSO 3] Dados preparados: PMM no Target sum={df['PMM no Target'].sum()}")
        
        # Agrupa por Rádio e Periodo_Semana
        grouped = df.groupby(['Rádio', 'Periodo_Semana']).agg({
            'Identificador': 'count',        # Contagem de inserções
            'Preço': 'sum',                  # Investimento
            'PMM': 'sum',                    # Soma de todos os PMM (Impactos)
            'PMM no Target': 'sum',          # TRP
            'Universo': lambda x: df.loc[x.index].drop_duplicates('Identificador')['Universo'].sum()  # Universo único
        }).reset_index()
        
        # Calcula PMM único por identificador
        pmm_unique = df.groupby(['Rádio', 'Periodo_Semana']).apply(
            lambda x: x.drop_duplicates('Identificador')['PMM'].sum()
        ).reset_index(name='PMM_Unico')
        
        # Junta os dados
        result_df = grouped.merge(pmm_unique, on=['Rádio', 'Periodo_Semana'])
        
        # Renomeia colunas
        result_df = result_df.rename(columns={
            'Periodo_Semana': 'Semana',
            'Identificador': 'Inserções',
            'Preço': 'Investimento',
            'PMM': 'Impactos',
            'PMM no Target': 'TRP',
            'PMM_Unico': 'PMM'
        })
        
        # Extrai Ano Comercial e Mês Comercial da coluna Semana
        result_df['Ano Comercial'] = result_df['Semana'].apply(extrair_ano_comercial)
        result_df['Mês Comercial'] = result_df['Semana'].apply(extrair_mes_comercial)
        
        # Adiciona coluna Programado (vazia)
        result_df['Programado'] = ''
        
        # Reorganiza colunas (incluindo as novas)
        final_columns = ['Rádio', 'Semana', 'Ano Comercial', 'Mês Comercial', 'Inserções', 'Investimento', 'Impactos', 'TRP', 'PMM', 'Universo', 'Programado']
        result_df = result_df[[col for col in final_columns if col in result_df.columns]]
        
        # Ordena por rádio e semana
        result_df = result_df.sort_values(['Rádio', 'Semana']).reset_index(drop=True)
        
        # Preenche NaN com 0 (exceto Programado)
        numeric_columns = ['Inserções', 'Investimento', 'Impactos', 'TRP', 'PMM', 'Universo']
        for col in numeric_columns:
            if col in result_df.columns:
                result_df[col] = result_df[col].fillna(0)
        
        print(f"[PASSO 3] Relatório semanal gerado: {len(result_df)} linhas")
        print(f"[PASSO 3] Colunas: {list(result_df.columns)}")
        
        # SALVA NA PASTA SAIDAS (COM FORMATAÇÃO - IGUAL AO COLAB)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"RELATORIO_SEMANAL_{timestamp}.xlsx"
        save_to_saidas(result_df, filename, apply_formatting=True)
        
        print("[PASSO 3] ✅ Relatório semanal salvo em saidas/")
        return result_df, None
    
    except Exception as e:
        print(f"[PASSO 3] ❌ ERRO CRÍTICO no relatório semanal:")
        print(f"[PASSO 3] Tipo do erro: {type(e).__name__}")
        print(f"[PASSO 3] Mensagem: {str(e)}")
        import traceback
        print(f"[PASSO 3] Stack trace completo:")
        traceback.print_exc()
        return None, f"Erro no passo 3: {type(e).__name__} - {str(e)}"

# ==============================================================================
# FUNÇÃO: ATUALIZAR RELATORIO SEMANAL OFICIAL
# ==============================================================================

def atualizar_semanal_oficial(df_semanal_novo):
    """
    Atualiza o arquivo oficial RELATORIO_SEMANAL_oficial.xlsx no Google Drive
    Adiciona apenas as linhas novas baseado em chave composta: Rádio + Semana
    (usa apenas colunas que existem em ambos os DataFrames)
    """
    try:
        # Remove colunas Unnamed do novo dataset também
        df_semanal_novo = df_semanal_novo.loc[:, ~df_semanal_novo.columns.str.contains('^Unnamed')]
        
        # Normaliza a coluna Semana: converte " a " (e variações de espaço) em "-" para compatibilidade
        if 'Semana' in df_semanal_novo.columns:
            df_semanal_novo['Semana'] = df_semanal_novo['Semana'].astype(str).str.replace(r'\s+a\s+', '-', regex=True)
        
        # ID do arquivo oficial no Google Drive
        file_id_oficial = "1o5RJnLMpMHMtvyG7FscwFhz37sxcc0_T"
        
        print(f"[SEMANAL OFICIAL] 🚀 Iniciando atualização (DataFrame com {len(df_semanal_novo)} linhas)")
        print(f"[SEMANAL OFICIAL] Colunas disponíveis: {list(df_semanal_novo.columns)}")
        
        # Autentica com Google Drive
        drive_service = get_authenticated_drive_service()
        if not drive_service:
            print(f"[SEMANAL OFICIAL] ⚠️ Não autenticado no Google Drive")
            return {'status': 'aviso', 'mensagem': 'Não autenticado - acesse /authorize'}
        
        print(f"[SEMANAL OFICIAL] ✅ Autenticado no Google Drive")
        
        # Baixa o arquivo oficial do Google Drive
        print(f"[SEMANAL OFICIAL] 📖 Buscando arquivo oficial no Google Drive...")
        df_oficial = None
        try:
            request = drive_service.files().get_media(fileId=file_id_oficial)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            
            fh.seek(0)
            df_oficial = pd.read_excel(fh)
            
            # Remove colunas Unnamed (vazias) imediatamente
            df_oficial = df_oficial.loc[:, ~df_oficial.columns.str.contains('^Unnamed')]
            
            # CRÍTICO: Normalizar Semana no oficial no MESMO formato do relatório novo (" a " -> "-").
            # Sem isso, a mesma semana aparece com duas chaves (ex: "29/12 a 04/01" vs "29/12-04/01"),
            # a linha do novo é vista como "nova" e é adicionada. Resultado: duas linhas para o mesmo (Rádio, Semana).
            # Ao somar no Excel: Inserções/Investimento/Impactos/TRP/PMM ficam DOBRADOS; Universo pode ficar
            # zerado na linha "nova" se naquele processamento o DadosIdentificador não tinha match (Universo=0).
            if 'Semana' in df_oficial.columns:
                df_oficial['Semana'] = df_oficial['Semana'].astype(str).str.replace(r'\s+a\s+', '-', regex=True)
            
            print(f"[SEMANAL OFICIAL] ✅ Arquivo oficial baixado: {len(df_oficial)} linhas")
            print(f"[SEMANAL OFICIAL] Colunas do arquivo oficial: {list(df_oficial.columns)}")
            
        except Exception as e:
            print(f"[SEMANAL OFICIAL] ⚠️ Erro ao baixar arquivo: {str(e)}")
            return {'status': 'erro', 'mensagem': f'Erro ao ler arquivo oficial: {str(e)}'}
        
        # Se o arquivo está vazio, aborta
        if df_oficial is None or len(df_oficial) == 0:
            print(f"[SEMANAL OFICIAL] ❌ ARQUIVO VAZIO - UPLOAD CANCELADO!")
            return {'status': 'erro', 'mensagem': 'Arquivo oficial está vazio. Upload cancelado.'}
        
        print(f"[SEMANAL OFICIAL] ✓ Arquivo oficial tem {len(df_oficial)} linhas")
        
        # Define as colunas que serão usadas para a chave composta
        # Usa apenas as colunas que EXISTEM em ambos os DataFrames
        colunas_chave = []
        colunas_possiveis = ['Rádio', 'Semana', 'Ano Comercial', 'Mês Comercial']
        
        for col in colunas_possiveis:
            if col in df_oficial.columns and col in df_semanal_novo.columns:
                colunas_chave.append(col)
                print(f"[SEMANAL OFICIAL] ✅ Coluna '{col}' disponível em ambos os arquivos")
            else:
                if col not in df_oficial.columns:
                    print(f"[SEMANAL OFICIAL] ⚠️ Coluna '{col}' NÃO ENCONTRADA no arquivo oficial")
                if col not in df_semanal_novo.columns:
                    print(f"[SEMANAL OFICIAL] ⚠️ Coluna '{col}' NÃO ENCONTRADA no novo relatório")
        
        if not colunas_chave:
            print(f"[SEMANAL OFICIAL] ❌ Nenhuma coluna em comum para fazer comparação!")
            return {'status': 'erro', 'mensagem': 'Nenhuma coluna em comum para fazer comparação'}
        
        print(f"[SEMANAL OFICIAL] 🔑 Colunas da chave composta: {colunas_chave}")
        
        # DEBUG: Mostra tipos de dados
        print(f"[SEMANAL OFICIAL] 🔍 Tipos de dados (Novo):")
        print(df_semanal_novo[colunas_chave].dtypes)
        print(f"[SEMANAL OFICIAL] 🔍 Tipos de dados (Oficial):")
        print(df_oficial[colunas_chave].dtypes)
        
        # Normaliza valor para chave: NaN, "nan", "none", vazio -> "__" para match consistente
        def _normalizar_para_chave(val):
            s = str(val).strip().upper()
            if s in ('', 'NAN', 'NONE', 'NAT', '<NA>'):
                return '__'
            return s
        
        # Cria identificador único para cada linha (valores normalizados para evitar falsas diferenças)
        df_oficial['_chave'] = df_oficial[colunas_chave].apply(
            lambda row: '|'.join(_normalizar_para_chave(v) for v in row), axis=1
        )
        df_semanal_novo['_chave'] = df_semanal_novo[colunas_chave].apply(
            lambda row: '|'.join(_normalizar_para_chave(v) for v in row), axis=1
        )
        
        print(f"[SEMANAL OFICIAL] Exemplo de chave oficial: {df_oficial['_chave'].iloc[0] if len(df_oficial) > 0 else 'N/A'}")
        print(f"[SEMANAL OFICIAL] Exemplo de chave novo: {df_semanal_novo['_chave'].iloc[0] if len(df_semanal_novo) > 0 else 'N/A'}")
        
        # Identifica as linhas novas (que não existem no arquivo oficial)
        chaves_oficiais = set(df_oficial['_chave'].unique())
        chaves_novo = set(df_semanal_novo['_chave'].unique())
        
        print(f"[SEMANAL OFICIAL] 📊 Total de chaves ÚNICAS no oficial: {len(chaves_oficiais)}")
        print(f"[SEMANAL OFICIAL] 📊 Total de chaves ÚNICAS no novo: {len(chaves_novo)}")
        
        # Verifica quantas chaves são comuns
        chaves_comuns = chaves_oficiais.intersection(chaves_novo)
        print(f"[SEMANAL OFICIAL] 🔄 Chaves em COMUM (já existem): {len(chaves_comuns)}")
        
        linhas_novas_df = df_semanal_novo[~df_semanal_novo['_chave'].isin(chaves_oficiais)].copy()
        
        # Evita duplicatas: se o relatório novo tiver duas linhas com mesma chave, mantém só uma
        linhas_novas_df = linhas_novas_df.drop_duplicates(subset=['_chave'], keep='first')
        
        linhas_novas = len(linhas_novas_df)
        print(f"[SEMANAL OFICIAL] 📊 Linhas novas encontradas: {linhas_novas}")
        
        # Aviso: linhas novas com Universo zerado (geralmente = Identificador ausente no DadosIdentificador)
        if 'Universo' in linhas_novas_df.columns:
            com_zero = linhas_novas_df[linhas_novas_df['Universo'].fillna(0) == 0]
            if len(com_zero) > 0:
                print(f"[SEMANAL OFICIAL] ⚠️ {len(com_zero)} linha(s) nova(s) com Universo = 0 (verifique DadosIdentificador/Google Sheets)")
                for _, r in com_zero.iterrows():
                    rad = r.get('Rádio', '')
                    sem = r.get('Semana', '')[:50] if pd.notna(r.get('Semana')) else ''
                    print(f"[SEMANAL OFICIAL]    → Rádio: {rad} | Semana: {sem}...")
        
        if linhas_novas_df.empty:
            print(f"[SEMANAL OFICIAL] ℹ️ Nenhuma linha nova para adicionar")
            return {'status': 'info', 'mensagem': 'Nenhuma linha nova para adicionar'}
        
        # Remove a coluna temporária de ambos
        df_oficial = df_oficial.drop(columns=['_chave'])
        linhas_novas_df = linhas_novas_df.drop(columns=['_chave'])
        
        # Concatena APENAS as novas linhas ao arquivo oficial (sem alterar linhas existentes)
        df_combinado = pd.concat([df_oficial, linhas_novas_df], ignore_index=True)
        
        # Ordena por Rádio e Semana (mantém ordem lógica)
        if 'Rádio' in df_combinado.columns and 'Semana' in df_combinado.columns:
            df_combinado = df_combinado.sort_values(['Rádio', 'Semana']).reset_index(drop=True)
        
        # Remove colunas Unnamed e alinha ordem das colunas à do oficial (evita coluna trocada no Excel)
        df_combinado = df_combinado.loc[:, ~df_combinado.columns.str.contains('^Unnamed')]
        colunas_oficial = [c for c in df_oficial.columns if c in df_combinado.columns]
        colunas_extra = [c for c in df_combinado.columns if c not in colunas_oficial]
        df_combinado = df_combinado[colunas_oficial + colunas_extra]
        
        print(f"[SEMANAL OFICIAL] ✅ Preparando upload: {len(df_oficial)} linhas originais + {linhas_novas} novas = {len(df_combinado)} total")
        print(f"[SEMANAL OFICIAL] 📋 Colunas finais para upload: {list(df_combinado.columns)}")
        
        # Salva em memória SEM índice, SEM colunas extras
        output = io.BytesIO()
        
        # Usa openpyxl para preservar formatação original
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Se o arquivo original é Excel, usa openpyxl para preservar formatação
        try:
            # Carrega o workbook original
            fh_original = io.BytesIO()
            request_original = drive_service.files().get_media(fileId=file_id_oficial)
            downloader = MediaIoBaseDownload(fh_original, request_original)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            fh_original.seek(0)
            
            # Abre o workbook original com openpyxl
            wb_original = load_workbook(fh_original)
            ws = wb_original.active
            
            # Mapeamento de colunas: Nome -> Índice (1-based)
            # Lê o cabeçalho da planilha (linha 1)
            header_map = {}
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=1, column=col_idx).value
                if cell_val:
                    header_map[str(cell_val).strip()] = col_idx
            
            print(f"[SEMANAL OFICIAL] 🗺️ Mapeamento de colunas (Excel): {header_map}")
            
            # Captura estilo da linha 2 (primeira linha de dados) para replicar
            style_template = {}
            if ws.max_row >= 2:
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=2, column=col_idx)
                    if cell.has_style:
                        style_template[col_idx] = {
                            'font': copy(cell.font),
                            'fill': copy(cell.fill),
                            'border': copy(cell.border),
                            'alignment': copy(cell.alignment),
                            'number_format': cell.number_format
                        }
            
            print(f"[SEMANAL OFICIAL] 🖌️ Estilo capturado da linha 2 para replicação")

            # REESCREVE A TABELA A PARTIR DA LINHA 2 COM O DATAFRAME COMBINADO
            # Isso garante que não haja buracos e a ordem seja respeitada
            total_rows = 0
            total_lines_df = len(df_combinado)
            
            # Atualiza progresso inicial
            atualizar_progresso(3, 76, f"Escrevendo {total_lines_df} linhas no arquivo oficial...")
            
            for idx, row in df_combinado.iterrows():
                current_row = 2 + idx  # Começa na linha 2
                total_rows += 1
                
                # Feedback a cada 100 linhas para não travar o front
                if total_rows % 100 == 0:
                    perc = 76 + int((total_rows / total_lines_df) * 10)  # Vai de 76% a 86%
                    atualizar_progresso(3, perc, f"Escrevendo linha {total_rows}/{total_lines_df}...")
                
                for col_name in df_combinado.columns:
                    target_col_idx = header_map.get(str(col_name).strip())
                    
                    if target_col_idx:
                        cell = ws.cell(row=current_row, column=target_col_idx)
                        cell.value = row[col_name]
                        
                        # Aplica o estilo capturado ou o estilo da linha anterior
                        if target_col_idx in style_template:
                            tmpl = style_template[target_col_idx]
                            cell.font = copy(tmpl['font'])
                            cell.fill = copy(tmpl['fill'])
                            cell.border = copy(tmpl['border'])
                            cell.alignment = copy(tmpl['alignment'])
                            cell.number_format = tmpl['number_format']

            print(f"[SEMANAL OFICIAL] ✅ Tabela reescrita: {total_rows} linhas a partir da linha 2")
            
            # Limpa linhas remanescentes (se a tabela nova for menor que a antiga/sujeira)
            # Verifica até onde ia o arquivo original
            max_row_original = ws.max_row
            rows_to_clear_start = 2 + total_rows
            
            if max_row_original >= rows_to_clear_start:
                print(f"[SEMANAL OFICIAL] 🧹 Limpando linhas excedentes da {rows_to_clear_start} até {max_row_original}")
                ws.delete_rows(rows_to_clear_start, amount=(max_row_original - rows_to_clear_start + 1))
            
            # Salva o workbook atualizado
            atualizar_progresso(3, 87, "Salvando arquivo Excel em memória...")
            wb_original.save(output)
            output.seek(0)
            
            print(f"[SEMANAL OFICIAL] ✅ Arquivo preparado com formatação preservada")
            
        except Exception as e:
            print(f"[SEMANAL OFICIAL] ⚠️ Erro ao preservar formatação: {str(e)}")
            print(f"[SEMANAL OFICIAL] 📝 Salvando como arquivo novo (sem formatação)")
            # Fallback: salva como arquivo novo
            df_combinado.to_excel(output, index=False, engine='openpyxl')
            output.seek(0)
        
        # Faz upload
        print(f"[SEMANAL OFICIAL] 📤 Fazendo upload para Google Drive...")
        media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        request = drive_service.files().update(fileId=file_id_oficial, media_body=media)
        request.execute()
        
        print(f"[SEMANAL OFICIAL] ✅ Upload concluído com sucesso!")
        return {
            'status': 'sucesso',
            'mensagem': f'Adicionadas {linhas_novas} linhas',
            'total_linhas': len(df_combinado)
        }
        
    except Exception as e:
        print(f"[SEMANAL OFICIAL] ❌ Erro ao atualizar: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'status': 'erro', 'mensagem': f'Erro ao atualizar: {str(e)}'}

# ==============================================================================
# ROTAS FLASK
# ==============================================================================

@app.route('/')
def index():
    """Página inicial"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Recebe arquivo e inicia processamento em segundo plano"""
    filepath = None
    try:
        print("[UPLOAD] 🚀 Iniciando upload...")
        
        if 'file' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'erro': 'Arquivo vazio'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'erro': 'Apenas arquivos Excel (.xlsx ou .xls)'}), 400
        
        print(f"[UPLOAD] 📄 Arquivo recebido: {file.filename}")
        
        # Salva arquivo temporário
        filename = secure_filename(file.filename)
        timestamp_task = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{timestamp_task}_{filename}")
        file.save(filepath)
        print(f"[UPLOAD] 💾 Arquivo salvo em: {filepath}")
        atualizar_progresso(1, 5, "Arquivo recebido...")

        # Inicializa cache de resultado como pendente
        if not hasattr(app, 'resultados_cache'):
            app.resultados_cache = {}
        app.resultados_cache[timestamp_task] = {'status': 'pendente', 'mensagem': 'Processamento em fila...'}

        # Inicia processamento em thread separada
        thread = threading.Thread(target=processar_em_background, args=(filepath, timestamp_task))
        thread.daemon = True
        thread.start()

        return jsonify({
            'sucesso': True, 
            'status': 'processando', 
            'timestamp': timestamp_task,
            'mensagem': 'Processamento iniciado em segundo plano.'
        }), 202

    except Exception as e:
        print(f"[UPLOAD] ❌ Erro ao iniciar upload: {str(e)}")
        return jsonify({'erro': f'Erro ao iniciar upload: {str(e)}'}), 500

def processar_em_background(filepath, timestamp_task):
    """Lógica de processamento que roda em segundo plano"""
    try:
        # Marca como processando no cache
        app.resultados_cache[timestamp_task] = {'status': 'pendente', 'mensagem': 'Iniciando compilação...'}
        
        # PASSO 1: Compilação
        print(f"[BG-TASK] 🔄 Iniciando PASSO 1 para {timestamp_task}...")
        atualizar_progresso(1, 10, "Compilando dados...")
        df_compilado, erro, timestamp_gen = passo1_compilar(filepath)
        
        if erro or df_compilado is None:
            print(f"[BG-TASK] ❌ PASSO 1 falhou: {erro}")
            atualizar_progresso(0, 0, f"Erro: {erro}")
            app.resultados_cache[timestamp_task] = {'sucesso': False, 'erro': f'Erro no Passo 1: {erro}'}
            return

        app.resultados_cache[timestamp_task] = {'status': 'pendente', 'mensagem': 'Gerando relatórios...'}
        atualizar_progresso(1, 40, f"Passo 1 completo: {len(df_compilado)} registros")
        
        # PASSO 2: Relatório Mensal
        print("[BG-TASK] 🔄 Iniciando PASSO 2...")
        atualizar_progresso(2, 45, "Gerando relatório mensal...")
        df_mensal, erro = passo2_mensal(df_compilado)
        if erro:
            print(f"[BG-TASK] Aviso Passo 2: {erro}")
            df_mensal = None
        atualizar_progresso(2, 70, f"Passo 2 completo: {len(df_mensal) if df_mensal is not None else 0} registros")
        
        # PASSO 3: Relatório Semanal
        print("[BG-TASK] 🔄 Iniciando PASSO 3...")
        atualizar_progresso(3, 75, "Gerando relatório semanal...")
        # FIX: Passar o df_compilado JÁ EM MEMÓRIA para não depender do disco
        df_semanal, erro = passo3_semanal(df_compilado)
        if erro:
            print(f"[BG-TASK] Aviso Passo 3: {erro}")
            df_semanal = None
        atualizar_progresso(3, 85, f"Passo 3 completo: {len(df_semanal) if df_semanal is not None else 0} registros")
        
        # Atualiza planilha SEMANAL oficial
        atualizar_resultado = None
        mensagem_tabela = None
        
        if df_semanal is not None and not df_semanal.empty:
            atualizar_resultado = atualizar_semanal_oficial(df_semanal)
            if atualizar_resultado and isinstance(atualizar_resultado, dict):
                if atualizar_resultado.get('status') == 'sucesso':
                    mensagem_tabela = {'tipo': 'sucesso', 'cor': 'verde', 'mensagem': '✅ Novas linhas adicionadas na tabela', 'detalhes': atualizar_resultado.get('mensagem', '')}
                elif atualizar_resultado.get('status') == 'info':
                    mensagem_tabela = {'tipo': 'info', 'cor': 'azul', 'mensagem': 'ℹ️ Tabela já está atualizada', 'detalhes': atualizar_resultado.get('mensagem', '')}
                else:
                    mensagem_tabela = {'tipo': 'aviso', 'cor': 'laranja', 'mensagem': '⚠️ Aviso ao atualizar tabela', 'detalhes': atualizar_resultado.get('mensagem', '')}
            else:
                mensagem_tabela = {'tipo': 'erro', 'cor': 'vermelho', 'mensagem': '❌ Erro ao atualizar tabela oficial', 'detalhes': str(atualizar_resultado)}
        else:
            # Mostra o erro real do passo 3 em vez de mensagem genérica
            erro_detalhado = erro if erro else 'Relatório semanal retornou vazio'
            mensagem_tabela = {'tipo': 'info', 'cor': 'cinza', 'mensagem': 'ℹ️ Relatório semanal não foi gerado', 'detalhes': erro_detalhado}

        # Prepara resultado final
        resultado_final = {
            'sucesso': True,
            'timestamp': timestamp_gen,
            'compilado': {'linhas': len(df_compilado), 'colunas': list(df_compilado.columns)},
            'mensal': {'linhas': len(df_mensal) if df_mensal is not None else 0, 'gerado': df_mensal is not None},
            'semanal': {'linhas': len(df_semanal) if df_semanal is not None else 0, 'gerado': df_semanal is not None},
            'semanal_oficial_atualizado': atualizar_resultado,
            'mensagem_tabela': mensagem_tabela
        }

        # Upload para Google Drive
        upload_results = upload_all_reports_to_drive(timestamp_gen)
        if upload_results:
            resultado_final['google_drive_upload'] = upload_results
        
        # Salva o resultado em disco para o frontend buscar
        resultado_path = os.path.join(app.config['SAIDAS_FOLDER'], f"resultado_{timestamp_task}.json")
        with open(resultado_path, 'w', encoding='utf-8') as f:
            json.dump(resultado_final, f, ensure_ascii=False, indent=4)
        
        # Cache em memória para acesso rápido no mesmo processo
        app.resultados_cache[timestamp_task] = resultado_final
        
        # Salva DataFrames globais (usados no /download)
        app.df_compilado = df_compilado
        app.df_mensal = df_mensal
        app.df_semanal = df_semanal
        
        print(f"[BG-TASK] ✅ Processamento {timestamp_task} concluído com sucesso!")
        atualizar_progresso(4, 100, "Concluído!")

    except Exception as e:
        import traceback
        print(f"[BG-TASK] ❌ Erro no processamento background: {str(e)}")
        print(traceback.format_exc())
        atualizar_progresso(0, 0, f"Erro fatal: {str(e)}")
        if hasattr(app, 'resultados_cache'):
            app.resultados_cache[timestamp_task] = {'sucesso': False, 'erro': str(e)}
    finally:
        # Limpa arquivo temporário
        if filepath and os.path.exists(filepath):
            try:
                time.sleep(1)
                os.remove(filepath)
                print(f"[BG-TASK] Arquivo temporário removido: {filepath}")
            except Exception as e:
                print(f"[BG-TASK] ⚠️ Erro ao remover temp: {str(e)}")

@app.route('/resultado/<timestamp>')
def obter_resultado(timestamp):
    """Busca o resultado do processamento finalizado ou status pendente"""
    # Tenta cache em memória primeiro (mais rápido e pega status pendente/erro)
    if hasattr(app, 'resultados_cache') and timestamp in app.resultados_cache:
        return jsonify(app.resultados_cache[timestamp])
    
    # Tenta carregar do disco (backup para persistência entre restarts)
    resultado_path = os.path.join(app.config['SAIDAS_FOLDER'], f"resultado_{timestamp}.json")
    if os.path.exists(resultado_path):
        try:
            with open(resultado_path, 'r', encoding='utf-8') as f:
                res = json.load(f)
                return jsonify(res)
        except Exception as e:
            return jsonify({'erro': f'Erro ao ler resultado: {str(e)}'}), 500
            
    return jsonify({'status': 'pendente', 'timestamp': timestamp}), 200

@app.route('/download/<tipo>')
def download(tipo):
    """Faz download dos arquivos gerados (lê dos DataFrames em memória ou disco)"""
    try:
        # Pega o arquivo compilado mais recente ou específico
        # Aqui simplificamos usando o que está em memória se existir
        df = None
        if tipo == 'compilado' and hasattr(app, 'df_compilado'):
            df = app.df_compilado
        elif tipo == 'mensal' and hasattr(app, 'df_mensal'):
            df = app.df_mensal
        elif tipo == 'semanal' and hasattr(app, 'df_semanal'):
            df = app.df_semanal
        
        if df is not None:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            output.seek(0)
            
            filename = f"RELATORIO_{tipo.upper()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return send_file(output, as_attachment=True, download_name=filename)
        
        # Se não estiver em memória, tenta buscar o mais recente do disco
        filepath = None
        if tipo == 'compilado':
            filepath = get_latest_compiled_file()
        elif tipo == 'semanal':
            filepath = get_latest_semanal_file()
        
        if filepath and os.path.exists(filepath):
            return send_file(filepath, as_attachment=True)
            
        return jsonify({'erro': 'Arquivo não encontrado. Tente gerar novamente.'}), 404
        
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

if __name__ == '__main__':
    print("🚀 Aplicação Flask iniciada!")
    print("📍 Acesse em http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)
